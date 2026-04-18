"""
SNTL 통합 물류 플랫폼 — Data Dictionary Excel 생성기
총 50개 테이블 | PostgreSQL 16 기준
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── 색상 팔레트 ─────────────────────────────────────────────
HEADER_BG   = "1F3864"   # 헤더 배경 (네이비)
HEADER_FG   = "FFFFFF"   # 헤더 글자
DOMAIN_COLORS = {
    "회원/인증":    "D9E1F2",
    "코드관리":     "E2EFDA",
    "오더관리":     "FFF2CC",
    "운송관리":     "FCE4D6",
    "창고관리":     "EDEDED",
    "Tracking":    "E8D5F5",
    "회계/청구":    "D5E8D4",
    "시스템관리":   "F8CECC",
    "기초정보관리": "DAE8FC",
}
ROW_ALT_BG = "F5F5F5"
PK_COLOR   = "FFEB9C"
FK_COLOR   = "DDEBF7"

THIN = Side(style="thin", color="AAAAAA")
THICK = Side(style="medium", color="888888")

def border_thin():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def border_thick():
    return Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

def h_cell(ws, row, col, value, bg=HEADER_BG, fg=HEADER_FG, bold=True, size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="맑은 고딕", bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border_thin()
    return c

def d_cell(ws, row, col, value, bg=None, bold=False, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="맑은 고딕", size=9, bold=bold)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border_thin()
    return c

# ─── 테이블 정의 데이터 ──────────────────────────────────────
# 형식: (No, 영문컬럼명, 한글컬럼명, 데이터타입, 길이/정밀도, PK, FK, NOT_NULL, 기본값, 설명)

TABLES = {
    "회원/인증": [
        {
            "name_en": "individual_member",
            "name_ko": "개인회원",
            "desc": "개인 사용자 회원 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","개인회원 고유 ID (자동증가)"),
                (2,"login_id","로그인ID","VARCHAR","50","-","UNIQUE","Y","-","로그인에 사용하는 고유 ID"),
                (3,"password","비밀번호","VARCHAR","255","-","-","Y","-","BCrypt 암호화된 비밀번호"),
                (4,"name","이름","VARCHAR","100","-","-","N","-","회원 이름"),
                (5,"birth_date","생년월일","DATE","-","-","-","N","-","생년월일"),
                (6,"phone","휴대폰번호","VARCHAR","20","-","-","N","-","휴대폰번호 (숫자만)"),
                (7,"email","이메일","VARCHAR","100","-","-","N","-","이메일 주소"),
                (8,"zipcode","우편번호","VARCHAR","10","-","-","N","-","우편번호"),
                (9,"address","주소","VARCHAR","255","-","-","N","-","주소 (시/도 시/군/구 읍/면/동)"),
                (10,"address_detail","상세주소","VARCHAR","255","-","-","N","-","상세주소"),
                (11,"grade","회원등급","VARCHAR","20","-","-","Y","BASIC","등급 (BASIC/SILVER/GOLD/VIP)"),
                (12,"balance","잔액","NUMERIC","15,2","-","-","Y","0","충전 잔액"),
                (13,"status","상태","VARCHAR","20","-","-","Y","ACTIVE","ACTIVE/INACTIVE/WITHDRAWN"),
                (14,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (15,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "corporate",
            "name_ko": "법인",
            "desc": "법인 회원 기본 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","법인 고유 ID"),
                (2,"login_id","로그인ID","VARCHAR","50","-","UNIQUE","Y","-","법인 로그인 ID"),
                (3,"corp_name","법인명","VARCHAR","200","-","-","Y","-","법인(회사) 명칭"),
                (4,"ceo_name","대표자명","VARCHAR","100","-","-","N","-","법인 대표자 이름"),
                (5,"business_number","사업자등록번호","VARCHAR","20","-","-","N","-","사업자등록번호"),
                (6,"phone","대표전화","VARCHAR","20","-","-","N","-","법인 대표 전화번호"),
                (7,"email","대표이메일","VARCHAR","100","-","-","N","-","법인 대표 이메일"),
                (8,"zipcode","우편번호","VARCHAR","10","-","-","N","-","법인 주소 우편번호"),
                (9,"address","주소","VARCHAR","255","-","-","N","-","법인 주소"),
                (10,"address_detail","상세주소","VARCHAR","255","-","-","N","-","법인 상세주소"),
                (11,"balance","잔액","NUMERIC","15,2","-","-","Y","0","충전 잔액"),
                (12,"approval_status","심사상태","VARCHAR","20","-","-","Y","PENDING","PENDING/APPROVED/REJECTED"),
                (13,"status","상태","VARCHAR","20","-","-","Y","INACTIVE","ACTIVE/INACTIVE/WITHDRAWN"),
                (14,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (15,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "corporate_attachment",
            "name_ko": "법인첨부파일",
            "desc": "법인 심사용 첨부파일 (사업자등록증 등), MinIO 저장",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","첨부파일 고유 ID"),
                (2,"corporate_id","법인ID","BIGINT","-","-","FK(corporate.id)","Y","-","소속 법인 ID"),
                (3,"file_name","파일명","VARCHAR","255","-","-","Y","-","원본 파일명"),
                (4,"file_path","파일경로","VARCHAR","500","-","-","Y","-","MinIO 저장 경로"),
                (5,"file_size","파일크기","BIGINT","-","-","-","N","-","파일 크기 (bytes)"),
                (6,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
        {
            "name_en": "corporate_manager",
            "name_ko": "법인관리자",
            "desc": "법인의 관리자 계정 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","법인관리자 고유 ID"),
                (2,"corporate_id","법인ID","BIGINT","-","-","FK(corporate.id)","Y","-","소속 법인 ID"),
                (3,"login_id","로그인ID","VARCHAR","50","-","UNIQUE","Y","-","법인관리자 로그인 ID"),
                (4,"password","비밀번호","VARCHAR","255","-","-","Y","-","BCrypt 암호화된 비밀번호"),
                (5,"name","이름","VARCHAR","100","-","-","N","-","법인관리자 이름"),
                (6,"birth_date","생년월일","DATE","-","-","-","N","-","생년월일"),
                (7,"phone","휴대폰번호","VARCHAR","20","-","-","N","-","휴대폰번호"),
                (8,"email","이메일","VARCHAR","100","-","-","N","-","이메일 주소"),
                (9,"zipcode","우편번호","VARCHAR","10","-","-","N","-","우편번호"),
                (10,"address","주소","VARCHAR","255","-","-","N","-","주소"),
                (11,"address_detail","상세주소","VARCHAR","255","-","-","N","-","상세주소"),
                (12,"status","상태","VARCHAR","20","-","-","Y","ACTIVE","ACTIVE/INACTIVE/WITHDRAWN"),
                (13,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (14,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "department",
            "name_ko": "부서",
            "desc": "법인 소속 부서 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","부서 고유 ID"),
                (2,"corporate_id","법인ID","BIGINT","-","-","FK(corporate.id)","Y","-","소속 법인 ID"),
                (3,"dept_name","부서명","VARCHAR","100","-","-","N","-","부서명"),
                (4,"status","상태","VARCHAR","20","-","-","Y","ACTIVE","ACTIVE/INACTIVE/WITHDRAWN"),
                (5,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (6,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "department_manager",
            "name_ko": "부서관리자",
            "desc": "부서 관리자 계정 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","부서관리자 고유 ID"),
                (2,"department_id","부서ID","BIGINT","-","-","FK(department.id)","Y","-","소속 부서 ID"),
                (3,"login_id","로그인ID","VARCHAR","50","-","UNIQUE","Y","-","부서관리자 로그인 ID"),
                (4,"password","비밀번호","VARCHAR","255","-","-","Y","-","BCrypt 암호화된 비밀번호"),
                (5,"name","이름","VARCHAR","100","-","-","N","-","부서관리자 이름"),
                (6,"birth_date","생년월일","DATE","-","-","-","N","-","생년월일"),
                (7,"phone","휴대폰번호","VARCHAR","20","-","-","N","-","휴대폰번호"),
                (8,"email","이메일","VARCHAR","100","-","-","N","-","이메일 주소"),
                (9,"zipcode","우편번호","VARCHAR","10","-","-","N","-","우편번호"),
                (10,"address","주소","VARCHAR","255","-","-","N","-","주소"),
                (11,"address_detail","상세주소","VARCHAR","255","-","-","N","-","상세주소"),
                (12,"status","상태","VARCHAR","20","-","-","Y","ACTIVE","ACTIVE/INACTIVE/WITHDRAWN"),
                (13,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (14,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "refresh_token",
            "name_ko": "리프레시토큰",
            "desc": "JWT Refresh Token 관리 (Redis 연계)",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"member_key","회원키","VARCHAR","100","-","-","Y","-","회원 구분 키 (예: IND_1 / CORP_2)"),
                (3,"token","토큰값","TEXT","-","-","-","Y","-","JWT Refresh Token 값"),
                (4,"expires_at","만료일시","TIMESTAMP","-","-","-","Y","-","토큰 만료일시"),
                (5,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
        {
            "name_en": "member_grade_policy",
            "name_ko": "회원등급정책",
            "desc": "개인회원 등급별 조건 및 할인율 정책",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"grade","등급","VARCHAR","20","-","-","Y","-","BASIC/SILVER/GOLD/VIP"),
                (3,"min_send_count","최소발송횟수","INTEGER","-","-","-","Y","-","등급 부여 최소 발송 횟수"),
                (4,"discount_rate","할인율","NUMERIC","5,2","-","-","Y","-","운송비 할인율 (%)"),
                (5,"description","설명","VARCHAR","500","-","-","N","-","등급 설명"),
                (6,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (7,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "prepaid_charge_request",
            "name_ko": "선불금충전요청",
            "desc": "회원 선불금 충전 요청 및 관리자 확인 이력",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","충전요청 고유 ID"),
                (2,"member_type","회원타입","VARCHAR","20","-","-","Y","-","INDIVIDUAL / CORPORATE"),
                (3,"member_id","회원ID","BIGINT","-","-","-","Y","-","충전 요청 회원 ID"),
                (4,"charge_amount","충전금액","NUMERIC","15,2","-","-","Y","-","요청 충전 금액"),
                (5,"deposit_date","입금일","DATE","-","-","-","N","-","실제 입금 날짜"),
                (6,"depositor_name","입금자명","VARCHAR","100","-","-","N","-","은행 입금자명"),
                (7,"status","처리상태","VARCHAR","20","-","-","Y","PENDING","PENDING / CONFIRMED / REJECTED"),
                (8,"confirmed_by","확인관리자ID","BIGINT","-","-","-","N","-","확인한 관리자 ID"),
                (9,"confirmed_at","확인일시","TIMESTAMP","-","-","-","N","-","관리자 확인 일시"),
                (10,"memo","메모","VARCHAR","500","-","-","N","-","처리 메모"),
                (11,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","요청 등록일시"),
                (12,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "prepaid_refund_request",
            "name_ko": "선불금환불요청",
            "desc": "회원 선불금 환불 요청 및 처리 이력",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","환불요청 고유 ID"),
                (2,"member_type","회원타입","VARCHAR","20","-","-","Y","-","INDIVIDUAL / CORPORATE"),
                (3,"member_id","회원ID","BIGINT","-","-","-","Y","-","환불 요청 회원 ID"),
                (4,"refund_amount","환불금액","NUMERIC","15,2","-","-","Y","-","요청 환불 금액"),
                (5,"bank_name","은행명","VARCHAR","100","-","-","Y","-","환불 수령 은행명"),
                (6,"account_number","계좌번호","VARCHAR","50","-","-","Y","-","환불 수령 계좌번호"),
                (7,"account_holder","예금주","VARCHAR","100","-","-","Y","-","계좌 예금주명"),
                (8,"status","처리상태","VARCHAR","20","-","-","Y","PENDING","PENDING / CONFIRMED / REJECTED"),
                (9,"processed_by","처리관리자ID","BIGINT","-","-","-","N","-","처리한 관리자 ID"),
                (10,"processed_at","처리일시","TIMESTAMP","-","-","-","N","-","처리 일시"),
                (11,"memo","메모","VARCHAR","500","-","-","N","-","처리 메모"),
                (12,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","요청 등록일시"),
                (13,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
    ],
    "코드관리": [
        {
            "name_en": "country_code",
            "name_ko": "국가코드",
            "desc": "ISO 3166 국가코드 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"country_code","국가코드","VARCHAR","3","-","UNIQUE","Y","-","ISO 3166-1 alpha-2/3"),
                (3,"country_name_ko","국가명(한글)","VARCHAR","100","-","-","Y","-","국가 한글 명칭"),
                (4,"country_name_en","국가명(영문)","VARCHAR","100","-","-","Y","-","국가 영문 명칭"),
                (5,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
            ]
        },
        {
            "name_en": "airport_code",
            "name_ko": "공항코드",
            "desc": "IATA 공항코드 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"airport_code","공항코드","VARCHAR","10","-","UNIQUE","Y","-","IATA 공항코드"),
                (3,"airport_name_ko","공항명(한글)","VARCHAR","200","-","-","N","-","공항 한글 명칭"),
                (4,"airport_name_en","공항명(영문)","VARCHAR","200","-","-","Y","-","공항 영문 명칭"),
                (5,"country_code","국가코드","VARCHAR","3","-","FK(country_code)","Y","-","소속 국가코드"),
                (6,"city","도시명","VARCHAR","100","-","-","N","-","공항 소재 도시명"),
                (7,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
            ]
        },
        {
            "name_en": "port_code",
            "name_ko": "항구코드",
            "desc": "UN/LOCODE 항구코드 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"port_code","항구코드","VARCHAR","10","-","UNIQUE","Y","-","UN/LOCODE 항구코드"),
                (3,"port_name_ko","항구명(한글)","VARCHAR","200","-","-","N","-","항구 한글 명칭"),
                (4,"port_name_en","항구명(영문)","VARCHAR","200","-","-","Y","-","항구 영문 명칭"),
                (5,"country_code","국가코드","VARCHAR","3","-","FK(country_code)","Y","-","소속 국가코드"),
                (6,"city","도시명","VARCHAR","100","-","-","N","-","항구 소재 도시명"),
                (7,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
            ]
        },
        {
            "name_en": "airline_code",
            "name_ko": "항공사코드",
            "desc": "IATA 항공사코드 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"airline_code","항공사코드","VARCHAR","10","-","UNIQUE","Y","-","IATA 항공사코드"),
                (3,"airline_name_ko","항공사명(한글)","VARCHAR","200","-","-","N","-","항공사 한글 명칭"),
                (4,"airline_name_en","항공사명(영문)","VARCHAR","200","-","-","Y","-","항공사 영문 명칭"),
                (5,"country_code","국가코드","VARCHAR","3","-","FK(country_code)","Y","-","소속 국가코드"),
                (6,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
            ]
        },
        {
            "name_en": "courier_company",
            "name_ko": "택배사",
            "desc": "택배사 기본 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"company_code","택배사코드","VARCHAR","10","-","UNIQUE","Y","-","택배사 식별 코드"),
                (3,"company_name_ko","택배사명(한글)","VARCHAR","200","-","-","Y","-","택배사 한글 명칭"),
                (4,"company_name_en","택배사명(영문)","VARCHAR","200","-","-","N","-","택배사 영문 명칭"),
                (5,"country_code","국가코드","VARCHAR","3","-","FK(country_code)","Y","-","소속 국가코드"),
                (6,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
            ]
        },
        {
            "name_en": "code_group",
            "name_ko": "코드그룹",
            "desc": "공통코드 그룹 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"group_code","그룹코드","VARCHAR","50","-","UNIQUE","Y","-","코드 그룹 식별 코드"),
                (3,"group_name","그룹명","VARCHAR","100","-","-","Y","-","코드 그룹 명칭"),
                (4,"description","설명","VARCHAR","500","-","-","N","-","코드 그룹 설명"),
                (5,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
            ]
        },
        {
            "name_en": "common_code",
            "name_ko": "공통코드",
            "desc": "시스템 공통코드 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"group_code","그룹코드","VARCHAR","50","-","FK(code_group)","Y","-","소속 코드 그룹"),
                (3,"code","코드값","VARCHAR","50","-","-","Y","-","공통 코드 값"),
                (4,"code_name","코드명","VARCHAR","100","-","-","Y","-","코드 명칭"),
                (5,"sort_order","정렬순서","INTEGER","-","-","-","Y","0","코드 표시 순서"),
                (6,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
            ]
        },
    ],
    "오더관리": [
        {
            "name_en": "orders",
            "name_ko": "오더",
            "desc": "배송 오더 기본 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","오더 고유 ID"),
                (2,"order_no","오더번호","VARCHAR","30","-","UNIQUE","Y","-","오더 번호 (ORD-YYYYMMDD-NNNNN)"),
                (3,"member_type","회원타입","VARCHAR","20","-","-","Y","-","INDIVIDUAL/CORPORATE/DEPT"),
                (4,"member_id","회원ID","BIGINT","-","-","Y","-","-","오더 등록 회원 ID"),
                (5,"order_status","오더상태","VARCHAR","20","-","-","Y","REGISTERED","REGISTERED/PACKED/WAREHOUSED/RELEASED/DELIVERED"),
                (6,"memo","메모","VARCHAR","500","-","-","N","-","오더 메모"),
                (7,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (8,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "shipper_info",
            "name_ko": "송하인정보",
            "desc": "오더 송하인(발송자) 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"order_id","오더ID","BIGINT","-","-","FK(orders.id)","Y","-","소속 오더 ID"),
                (3,"name","이름","VARCHAR","100","-","-","Y","-","송하인 이름"),
                (4,"phone","전화번호","VARCHAR","20","-","-","N","-","송하인 연락처"),
                (5,"email","이메일","VARCHAR","100","-","-","N","-","송하인 이메일"),
                (6,"zipcode","우편번호","VARCHAR","10","-","-","N","-","우편번호"),
                (7,"address","주소","VARCHAR","255","-","-","N","-","주소"),
                (8,"address_detail","상세주소","VARCHAR","255","-","-","N","-","상세주소"),
                (9,"country_code","국가코드","VARCHAR","3","-","FK(country_code)","Y","-","발송 국가 코드"),
            ]
        },
        {
            "name_en": "consignee_info",
            "name_ko": "수하인정보",
            "desc": "오더 수하인(수취인) 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"order_id","오더ID","BIGINT","-","-","FK(orders.id)","Y","-","소속 오더 ID"),
                (3,"name","이름","VARCHAR","100","-","-","Y","-","수하인 이름"),
                (4,"phone","전화번호","VARCHAR","20","-","-","N","-","수하인 연락처"),
                (5,"email","이메일","VARCHAR","100","-","-","N","-","수하인 이메일"),
                (6,"zipcode","우편번호","VARCHAR","10","-","-","N","-","우편번호"),
                (7,"address","주소","VARCHAR","255","-","-","N","-","주소"),
                (8,"address_detail","상세주소","VARCHAR","255","-","-","N","-","상세주소"),
                (9,"country_code","국가코드","VARCHAR","3","-","FK(country_code)","Y","-","수취 국가 코드"),
                (10,"passport_no","여권번호","VARCHAR","20","-","-","N","-","수취인 여권번호 (통관 목적)"),
            ]
        },
        {
            "name_en": "cargo_info",
            "name_ko": "화물정보",
            "desc": "오더 화물 상세 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"order_id","오더ID","BIGINT","-","-","FK(orders.id)","Y","-","소속 오더 ID"),
                (3,"item_name_ko","품명(한글)","VARCHAR","200","-","-","Y","-","화물 품명 (한글)"),
                (4,"item_name_en","품명(영문)","VARCHAR","200","-","-","N","-","화물 품명 (영문)"),
                (5,"hs_code","HS코드","VARCHAR","20","-","-","N","-","HS 관세 분류 코드"),
                (6,"quantity","수량","INTEGER","-","-","-","Y","-","화물 수량"),
                (7,"weight_kg","중량(kg)","NUMERIC","10,3","-","-","Y","-","화물 총 중량 (kg)"),
                (8,"dim_x","가로(cm)","NUMERIC","10,2","-","-","N","-","박스 가로 크기"),
                (9,"dim_y","세로(cm)","NUMERIC","10,2","-","-","N","-","박스 세로 크기"),
                (10,"dim_z","높이(cm)","NUMERIC","10,2","-","-","N","-","박스 높이 크기"),
                (11,"declared_value","신고가격","NUMERIC","15,2","-","-","N","-","세관 신고 가격 (USD)"),
                (12,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
        {
            "name_en": "order_service",
            "name_ko": "오더서비스",
            "desc": "오더에 신청된 서비스 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"order_id","오더ID","BIGINT","-","-","FK(orders.id)","Y","-","소속 오더 ID"),
                (3,"service_type","서비스구분","VARCHAR","10","-","-","Y","-","AIR/SEA/CIR/CCL"),
                (4,"extra_charge","추가요금","NUMERIC","15,1","-","-","N","0","오더별 수동 추가 요금"),
                (5,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (6,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
    ],
    "운송관리": [
        {
            "name_en": "master_order",
            "name_ko": "마스터오더",
            "desc": "오더들을 묶어 운송 단위로 관리하는 마스터오더",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","마스터오더 고유 ID"),
                (2,"master_order_no","마스터오더번호","VARCHAR","30","-","UNIQUE","Y","-","마스터오더 번호 (MO-YYYYMMDD-NNNNN)"),
                (3,"service_type","서비스구분","VARCHAR","10","-","-","Y","-","AIR/SEA/CIR"),
                (4,"master_status","마스터오더상태","VARCHAR","20","-","-","Y","CREATED","CREATED/WAREHOUSED/RELEASED/IN_TRANSIT/DELIVERED"),
                (5,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (6,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "transport_schedule",
            "name_ko": "운항스케줄",
            "desc": "항공/해운 운항 스케줄 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"master_order_id","마스터오더ID","BIGINT","-","-","FK(master_order.id)","Y","-","소속 마스터오더 ID"),
                (3,"carrier_id","운송사ID","BIGINT","-","-","-","Y","-","항공사/선사 ID"),
                (4,"departure_code","출발코드","VARCHAR","10","-","-","N","-","출발 공항 또는 항구 코드"),
                (5,"arrival_code","도착코드","VARCHAR","10","-","-","N","-","도착 공항 또는 항구 코드"),
                (6,"departure_dt","출발일시","TIMESTAMP","-","-","-","N","-","출발 일시"),
                (7,"arrival_dt","도착일시","TIMESTAMP","-","-","-","N","-","도착 일시"),
                (8,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
    ],
    "창고관리": [
        {
            "name_en": "warehouse_receipt",
            "name_ko": "입고",
            "desc": "창고 입고 처리 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","입고 고유 ID"),
                (2,"order_id","오더ID","BIGINT","-","-","FK(orders.id)","Y","-","소속 오더 ID"),
                (3,"barcode","바코드","VARCHAR","50","-","-","N","-","입고 바코드"),
                (4,"receipt_dt","입고일시","TIMESTAMP","-","-","-","Y","NOW()","실제 입고 일시"),
                (5,"worker_id","처리자ID","BIGINT","-","-","-","N","-","입고 처리자 ID"),
                (6,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
        {
            "name_en": "warehouse_release",
            "name_ko": "출고",
            "desc": "창고 출고 처리 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","출고 고유 ID"),
                (2,"master_order_id","마스터오더ID","BIGINT","-","-","FK(master_order.id)","Y","-","소속 마스터오더 ID"),
                (3,"release_dt","출고일시","TIMESTAMP","-","-","-","Y","NOW()","실제 출고 일시"),
                (4,"worker_id","처리자ID","BIGINT","-","-","-","N","-","출고 처리자 ID"),
                (5,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
    ],
    "Tracking": [
        {
            "name_en": "tracking_info",
            "name_ko": "운송Tracking",
            "desc": "AIR/SEA/CIR 운송 추적 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"master_order_id","마스터오더ID","BIGINT","-","-","FK(master_order.id)","Y","-","소속 마스터오더 ID"),
                (3,"status_code","상태코드","VARCHAR","50","-","-","Y","-","외부 Tracking 상태 코드"),
                (4,"status_desc","상태설명","VARCHAR","200","-","-","N","-","상태 설명"),
                (5,"location","위치","VARCHAR","200","-","-","N","-","현재 위치"),
                (6,"event_dt","이벤트일시","TIMESTAMP","-","-","-","N","-","상태 발생 일시"),
                (7,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
        {
            "name_en": "customs_declaration",
            "name_ko": "통관신고",
            "desc": "통관 신고 및 진행 상태 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"order_id","오더ID","BIGINT","-","-","FK(orders.id)","Y","-","소속 오더 ID"),
                (3,"declaration_no","신고번호","VARCHAR","50","-","-","N","-","세관 신고 번호"),
                (4,"declaration_status","신고상태","VARCHAR","20","-","-","Y","PENDING","PENDING/ACCEPTED/CLEARED/REJECTED"),
                (5,"customs_broker_id","통관사ID","BIGINT","-","-","FK(customs_broker.id)","N","-","담당 통관사 ID"),
                (6,"declared_at","신고일시","TIMESTAMP","-","-","-","N","-","통관 신고 일시"),
                (7,"cleared_at","통관완료일시","TIMESTAMP","-","-","-","N","-","통관 완료 일시"),
                (8,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (9,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
    ],
    "회계/청구": [
        {
            "name_en": "invoice",
            "name_ko": "청구서",
            "desc": "운송비 청구서 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"invoice_no","청구서번호","VARCHAR","30","-","UNIQUE","Y","-","청구서 고유 번호"),
                (3,"member_type","회원타입","VARCHAR","20","-","-","Y","-","INDIVIDUAL/CORPORATE/DEPT"),
                (4,"member_id","회원ID","BIGINT","-","-","Y","-","-","청구 대상 회원 ID"),
                (5,"total_amount","합계금액","NUMERIC","15,2","-","-","Y","-","청구서 합계 금액"),
                (6,"invoice_status","청구상태","VARCHAR","20","-","-","Y","UNPAID","UNPAID/PARTIAL/PAID"),
                (7,"issued_dt","발행일","DATE","-","-","-","Y","-","청구서 발행일"),
                (8,"due_dt","납부기한","DATE","-","-","-","N","-","납부 기한"),
                (9,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (10,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "invoice_item",
            "name_ko": "청구항목",
            "desc": "청구서 내 오더별 청구 항목",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"invoice_id","청구서ID","BIGINT","-","-","FK(invoice.id)","Y","-","소속 청구서 ID"),
                (3,"order_id","오더ID","BIGINT","-","-","FK(orders.id)","Y","-","청구 대상 오더 ID"),
                (4,"service_type","서비스구분","VARCHAR","10","-","-","Y","-","AIR/SEA/CIR/CCL"),
                (5,"base_cost","기본운임","NUMERIC","15,2","-","-","Y","-","기본 운송 원가"),
                (6,"extra_charge","추가요금","NUMERIC","15,1","-","-","Y","0","오더별 추가 요금"),
                (7,"discount_amount","할인금액","NUMERIC","15,2","-","-","Y","0","적용 할인 금액"),
                (8,"final_amount","최종금액","NUMERIC","15,2","-","-","Y","-","최종 청구 금액"),
                (9,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
        {
            "name_en": "payment",
            "name_ko": "입금",
            "desc": "청구서 입금 처리 이력",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"invoice_id","청구서ID","BIGINT","-","-","FK(invoice.id)","Y","-","소속 청구서 ID"),
                (3,"payment_amount","입금금액","NUMERIC","15,2","-","-","Y","-","입금 금액"),
                (4,"payment_dt","입금일시","TIMESTAMP","-","-","-","Y","-","입금 일시"),
                (5,"payment_method","입금방법","VARCHAR","20","-","-","N","-","BANK/CARD/PREPAID"),
                (6,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
        {
            "name_en": "transport_cost_master",
            "name_ko": "운송원가마스터",
            "desc": "서비스/운송사/구간별 운송 원가 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"service_type","서비스구분","VARCHAR","10","-","-","Y","-","AIR/SEA/CIR"),
                (3,"carrier_id","운송사ID","BIGINT","-","-","-","Y","-","항공사/선사/택배사 ID"),
                (4,"departure_code","출발코드","VARCHAR","10","-","-","N","-","출발 공항 또는 항구 코드"),
                (5,"arrival_code","도착코드","VARCHAR","10","-","-","N","-","도착 공항 또는 항구 코드"),
                (6,"weight_from","중량구간시작","NUMERIC","10,3","-","-","N","-","중량 구간 시작값 (kg)"),
                (7,"weight_to","중량구간끝","NUMERIC","10,3","-","-","N","-","중량 구간 끝값 (kg)"),
                (8,"unit_cost","단위원가","NUMERIC","15,4","-","-","Y","-","단위당 운송 원가"),
                (9,"cost_unit","원가단위","VARCHAR","20","-","-","N","-","KG/CBM"),
                (10,"effective_from","적용시작일","DATE","-","-","-","Y","-","원가 적용 시작일"),
                (11,"effective_to","적용종료일","DATE","-","-","-","N","-","원가 적용 종료일 (NULL=현재적용중)"),
                (12,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (13,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "member_discount",
            "name_ko": "회원별할인율",
            "desc": "회원별 서비스 할인율 및 영업이익율",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"member_type","회원타입","VARCHAR","20","-","-","Y","-","INDIVIDUAL/CORPORATE/DEPT"),
                (3,"member_id","회원ID","BIGINT","-","-","-","Y","-","대상 회원 ID"),
                (4,"service_type","서비스구분","VARCHAR","10","-","-","Y","-","AIR/SEA/CIR/CCL"),
                (5,"discount_rate","할인율","NUMERIC","5,2","-","-","Y","-","회원 할인율 (%)"),
                (6,"profit_rate","영업이익율","NUMERIC","5,2","-","-","Y","-","영업 이익율 (%)"),
                (7,"effective_from","적용시작일","DATE","-","-","-","Y","-","적용 시작일"),
                (8,"effective_to","적용종료일","DATE","-","-","-","N","-","적용 종료일"),
                (9,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (10,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "tax_invoice",
            "name_ko": "세금계산서",
            "desc": "청구서 기반 세금계산서 정보",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"invoice_id","청구서ID","BIGINT","-","-","FK(invoice.id)","Y","-","소속 청구서 ID"),
                (3,"tax_invoice_no","세금계산서번호","VARCHAR","50","-","UNIQUE","Y","-","세금계산서 고유번호"),
                (4,"issue_dt","발행일","DATE","-","-","-","Y","-","세금계산서 발행일"),
                (5,"supply_amount","공급가액","NUMERIC","15,2","-","-","Y","-","공급가액"),
                (6,"tax_amount","세액","NUMERIC","15,2","-","-","Y","-","부가세액"),
                (7,"total_amount","합계금액","NUMERIC","15,2","-","-","Y","-","합계금액"),
                (8,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
    ],
    "시스템관리": [
        {
            "name_en": "notification",
            "name_ko": "알림",
            "desc": "시스템 알림 발송 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","알림 고유 ID"),
                (2,"title","제목","VARCHAR","200","-","-","Y","-","알림 제목"),
                (3,"content","내용","TEXT","-","-","-","Y","-","알림 내용"),
                (4,"target_type","대상타입","VARCHAR","20","-","-","N","-","ALL/INDIVIDUAL/CORPORATE"),
                (5,"target_id","대상ID","BIGINT","-","-","-","N","-","특정 대상 회원 ID"),
                (6,"send_status","발송상태","VARCHAR","20","-","-","Y","PENDING","PENDING/SENT/FAILED"),
                (7,"sent_dt","발송일시","TIMESTAMP","-","-","-","N","-","실제 발송 일시"),
                (8,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
        {
            "name_en": "voc",
            "name_ko": "VOC",
            "desc": "오더별 고객 불만/문의(VOC) 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","VOC 고유 ID"),
                (2,"order_id","오더ID","BIGINT","-","-","FK(orders.id)","Y","-","관련 오더 ID"),
                (3,"member_type","회원타입","VARCHAR","20","-","-","Y","-","문의 회원 타입"),
                (4,"member_id","회원ID","BIGINT","-","-","-","Y","-","문의 회원 ID"),
                (5,"content","문의내용","TEXT","-","-","-","Y","-","VOC 문의 내용"),
                (6,"reply","답변","TEXT","-","-","-","N","-","담당자 답변 내용"),
                (7,"status","처리상태","VARCHAR","20","-","-","Y","OPEN","OPEN/IN_PROGRESS/CLOSED"),
                (8,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (9,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "menu",
            "name_ko": "메뉴",
            "desc": "시스템 메뉴 구조 관리 (계층형)",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","메뉴 고유 ID"),
                (2,"parent_id","상위메뉴ID","BIGINT","-","-","FK(menu.id)","N","-","상위 메뉴 ID (NULL=최상위)"),
                (3,"menu_name","메뉴명","VARCHAR","100","-","-","Y","-","메뉴 명칭"),
                (4,"menu_url","메뉴URL","VARCHAR","200","-","-","N","-","메뉴 링크 URL"),
                (5,"sort_order","정렬순서","INTEGER","-","-","-","Y","0","메뉴 표시 순서"),
                (6,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
                (7,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (8,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "role",
            "name_ko": "권한",
            "desc": "시스템 권한(Role) 정의",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","권한 고유 ID"),
                (2,"role_code","권한코드","VARCHAR","50","-","UNIQUE","Y","-","권한 식별 코드"),
                (3,"role_name","권한명","VARCHAR","100","-","-","Y","-","권한 명칭"),
                (4,"description","설명","VARCHAR","500","-","-","N","-","권한 설명"),
            ]
        },
        {
            "name_en": "menu_role",
            "name_ko": "메뉴권한매핑",
            "desc": "메뉴와 권한 매핑 테이블",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"menu_id","메뉴ID","BIGINT","-","-","FK(menu.id)","Y","-","메뉴 ID"),
                (3,"role_id","권한ID","BIGINT","-","-","FK(role.id)","Y","-","권한 ID"),
            ]
        },
    ],
    "기초정보관리": [
        {
            "name_en": "individual_fare_rate",
            "name_ko": "개인회원운임요율",
            "desc": "개인회원 등급별 운임 할인/적용 요율",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"member_grade","회원등급","VARCHAR","20","-","-","Y","-","BASIC / SILVER / GOLD / VIP"),
                (3,"service_type","서비스구분","VARCHAR","10","-","-","Y","-","AIR / SEA / CIR / CCL"),
                (4,"fare_rate","운임요율","NUMERIC","5,2","-","-","Y","-","적용 운임 요율 (%)"),
                (5,"effective_from","적용시작일","DATE","-","-","-","Y","-","요율 적용 시작일"),
                (6,"effective_to","적용종료일","DATE","-","-","-","N","-","요율 적용 종료일 (NULL=현재적용중)"),
                (7,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (8,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "corporate_fare_rate",
            "name_ko": "법인회원운임요율",
            "desc": "법인별 맞춤 운임 적용 요율",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"corporate_id","법인ID","BIGINT","-","-","FK(corporate.id)","Y","-","대상 법인 ID"),
                (3,"service_type","서비스구분","VARCHAR","10","-","-","Y","-","AIR / SEA / CIR / CCL"),
                (4,"fare_rate","운임요율","NUMERIC","5,2","-","-","Y","-","적용 운임 요율 (%)"),
                (5,"effective_from","적용시작일","DATE","-","-","-","Y","-","요율 적용 시작일"),
                (6,"effective_to","적용종료일","DATE","-","-","-","N","-","요율 적용 종료일 (NULL=현재적용중)"),
                (7,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (8,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "exchange_rate",
            "name_ko": "환율",
            "desc": "서울외국환중개소 API 연계 환율 이력 저장",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"currency_code","통화코드","VARCHAR","10","-","-","Y","-","ISO 4217 통화코드 (예: USD, EUR)"),
                (3,"currency_name","통화명","VARCHAR","100","-","-","Y","-","통화 명칭"),
                (4,"base_rate","기준환율","NUMERIC","15,4","-","-","Y","-","원화 기준 환율 (KRW/외화)"),
                (5,"rate_date","기준일","DATE","-","-","-","Y","-","환율 기준 날짜 (영업일 오전 8시 기준)"),
                (6,"source","출처","VARCHAR","50","-","-","N","SEOUL_FX","환율 출처 (서울외국환중개소)"),
                (7,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
            ]
        },
        {
            "name_en": "air_transport",
            "name_ko": "항공운송수단",
            "desc": "항공 운송 구간 및 편명 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"origin_country_code","출발국가코드","VARCHAR","3","-","FK(country_code)","Y","-","출발 국가 코드"),
                (3,"origin_airport_code","출발공항코드","VARCHAR","10","-","FK(airport_code)","Y","-","출발 공항 코드"),
                (4,"dest_country_code","도착국가코드","VARCHAR","3","-","FK(country_code)","Y","-","도착 국가 코드"),
                (5,"dest_airport_code","도착공항코드","VARCHAR","10","-","FK(airport_code)","Y","-","도착 공항 코드"),
                (6,"flight_name","운송편명","VARCHAR","200","-","-","Y","-","항공 편명 또는 운항사명"),
                (7,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
                (8,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (9,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "air_transport_cost",
            "name_ko": "항공운송원가",
            "desc": "항공 운송수단별 부피/중량 기준 운송원가",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"air_transport_id","항공운송수단ID","BIGINT","-","-","FK(air_transport.id)","Y","-","소속 항공 운송수단 ID"),
                (3,"cost_type","원가구분","VARCHAR","10","-","-","Y","-","VOLUME (부피기준) / WEIGHT (중량기준)"),
                (4,"dim_x","가로(cm)","NUMERIC","10,2","-","-","N","-","부피기준 가로 (VOLUME 시 필수)"),
                (5,"dim_y","세로(cm)","NUMERIC","10,2","-","-","N","-","부피기준 세로 (VOLUME 시 필수)"),
                (6,"dim_z","높이(cm)","NUMERIC","10,2","-","-","N","-","부피기준 높이 (VOLUME 시 필수)"),
                (7,"weight_kg","중량(kg)","NUMERIC","10,3","-","-","N","-","중량기준 중량 (WEIGHT 시 필수)"),
                (8,"cost_usd","원가(USD)","NUMERIC","15,4","-","-","Y","-","단위당 운송 원가 (USD)"),
                (9,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (10,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "sea_transport",
            "name_ko": "해운운송수단",
            "desc": "해운 운송 구간 및 선사 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"origin_country_code","출발국가코드","VARCHAR","3","-","FK(country_code)","Y","-","출발 국가 코드"),
                (3,"origin_port_code","출발항구코드","VARCHAR","10","-","FK(port_code)","Y","-","출발 항구 코드"),
                (4,"dest_country_code","도착국가코드","VARCHAR","3","-","FK(country_code)","Y","-","도착 국가 코드"),
                (5,"dest_port_code","도착항구코드","VARCHAR","10","-","FK(port_code)","Y","-","도착 항구 코드"),
                (6,"vessel_name","운송편명","VARCHAR","200","-","-","Y","-","선사명 또는 선박편명"),
                (7,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
                (8,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (9,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "sea_transport_cost",
            "name_ko": "해운운송원가",
            "desc": "해운 운송수단별 부피/중량 기준 운송원가",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"sea_transport_id","해운운송수단ID","BIGINT","-","-","FK(sea_transport.id)","Y","-","소속 해운 운송수단 ID"),
                (3,"cost_type","원가구분","VARCHAR","10","-","-","Y","-","VOLUME (부피기준) / WEIGHT (중량기준)"),
                (4,"dim_x","가로(cm)","NUMERIC","10,2","-","-","N","-","부피기준 가로 (VOLUME 시 필수)"),
                (5,"dim_y","세로(cm)","NUMERIC","10,2","-","-","N","-","부피기준 세로 (VOLUME 시 필수)"),
                (6,"dim_z","높이(cm)","NUMERIC","10,2","-","-","N","-","부피기준 높이 (VOLUME 시 필수)"),
                (7,"weight_kg","중량(kg)","NUMERIC","10,3","-","-","N","-","중량기준 중량 (WEIGHT 시 필수)"),
                (8,"cost_usd","원가(USD)","NUMERIC","15,4","-","-","Y","-","단위당 운송 원가 (USD)"),
                (9,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (10,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "customs_broker",
            "name_ko": "통관사",
            "desc": "통관사 정보 및 API 연동 구분 관리",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"country_code","국가코드","VARCHAR","3","-","FK(country_code)","Y","-","통관 대상 국가 코드"),
                (3,"service_type","서비스구분","VARCHAR","10","-","-","Y","-","AIR / SEA"),
                (4,"entry_code","통관입력코드","VARCHAR","10","-","-","N","-","세관 시스템 입력 코드"),
                (5,"broker_name","통관사명","VARCHAR","200","-","-","Y","-","통관사 명칭"),
                (6,"api_integration","API연동구분","CHAR","3","-","-","Y","OFF","ON / OFF"),
                (7,"use_yn","사용여부","CHAR","1","-","-","Y","Y","사용여부 (Y/N)"),
                (8,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (9,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "customs_broker_cost",
            "name_ko": "통관원가",
            "desc": "통관사별 부피/중량 기준 통관원가",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"customs_broker_id","통관사ID","BIGINT","-","-","FK(customs_broker.id)","Y","-","소속 통관사 ID"),
                (3,"cost_type","원가구분","VARCHAR","10","-","-","Y","-","VOLUME (부피기준) / WEIGHT (중량기준)"),
                (4,"dim_x","가로(cm)","NUMERIC","10,2","-","-","N","-","부피기준 가로 (VOLUME 시 필수)"),
                (5,"dim_y","세로(cm)","NUMERIC","10,2","-","-","N","-","부피기준 세로 (VOLUME 시 필수)"),
                (6,"dim_z","높이(cm)","NUMERIC","10,2","-","-","N","-","부피기준 높이 (VOLUME 시 필수)"),
                (7,"weight_kg","중량(kg)","NUMERIC","10,3","-","-","N","-","중량기준 중량 (WEIGHT 시 필수)"),
                (8,"cost_usd","원가(USD)","NUMERIC","15,4","-","-","Y","-","단위당 통관 원가 (USD)"),
                (9,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (10,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "courier_transport_cost",
            "name_ko": "택배사운송원가",
            "desc": "택배사별 부피/중량 기준 운송원가",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"courier_id","택배사ID","BIGINT","-","-","FK(courier_company.id)","Y","-","소속 택배사 ID"),
                (3,"cost_type","원가구분","VARCHAR","10","-","-","Y","-","VOLUME (부피기준) / WEIGHT (중량기준)"),
                (4,"dim_x","가로(cm)","NUMERIC","10,2","-","-","N","-","부피기준 가로 (VOLUME 시 필수)"),
                (5,"dim_y","세로(cm)","NUMERIC","10,2","-","-","N","-","부피기준 세로 (VOLUME 시 필수)"),
                (6,"dim_z","높이(cm)","NUMERIC","10,2","-","-","N","-","부피기준 높이 (VOLUME 시 필수)"),
                (7,"weight_kg","중량(kg)","NUMERIC","10,3","-","-","N","-","중량기준 중량 (WEIGHT 시 필수)"),
                (8,"cost_usd","원가(USD)","NUMERIC","15,4","-","-","Y","-","단위당 운송 원가 (USD)"),
                (9,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (10,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
        {
            "name_en": "courier_waybill",
            "name_ko": "택배배송장",
            "desc": "택배사별 배송장 양식 파일 관리 (MinIO 저장)",
            "cols": [
                (1,"id","고유ID","BIGSERIAL","-","PK","-","Y","-","고유 ID"),
                (2,"courier_id","택배사ID","BIGINT","-","-","FK(courier_company.id)","Y","-","소속 택배사 ID"),
                (3,"country_code","국가코드","VARCHAR","3","-","FK(country_code)","Y","-","적용 국가 코드"),
                (4,"file_name","파일명","VARCHAR","255","-","-","Y","-","원본 파일명"),
                (5,"file_path","파일경로","VARCHAR","500","-","-","Y","-","MinIO 저장 경로"),
                (6,"file_size","파일크기","BIGINT","-","-","-","N","-","파일 크기 (bytes)"),
                (7,"description","설명","VARCHAR","500","-","-","N","-","배송장 양식 설명"),
                (8,"created_at","등록일시","TIMESTAMP","-","-","-","Y","NOW()","등록일시"),
                (9,"updated_at","수정일시","TIMESTAMP","-","-","-","Y","NOW()","최종수정일시"),
            ]
        },
    ],
}

# ─── 목차 데이터 ─────────────────────────────────────────────
INDEX_DATA = [
    (1,"individual_member","개인회원","회원/인증"),
    (2,"corporate","법인","회원/인증"),
    (3,"corporate_attachment","법인첨부파일","회원/인증"),
    (4,"corporate_manager","법인관리자","회원/인증"),
    (5,"department","부서","회원/인증"),
    (6,"department_manager","부서관리자","회원/인증"),
    (7,"refresh_token","리프레시토큰","회원/인증"),
    (8,"member_grade_policy","회원등급정책","회원/인증"),
    (9,"prepaid_charge_request","선불금충전요청","회원/인증"),
    (10,"prepaid_refund_request","선불금환불요청","회원/인증"),
    (11,"country_code","국가코드","코드관리"),
    (12,"airport_code","공항코드","코드관리"),
    (13,"port_code","항구코드","코드관리"),
    (14,"airline_code","항공사코드","코드관리"),
    (15,"courier_company","택배사","코드관리"),
    (16,"code_group","코드그룹","코드관리"),
    (17,"common_code","공통코드","코드관리"),
    (18,"orders","오더","오더관리"),
    (19,"shipper_info","송하인정보","오더관리"),
    (20,"consignee_info","수하인정보","오더관리"),
    (21,"cargo_info","화물정보","오더관리"),
    (22,"order_service","오더서비스","오더관리"),
    (23,"master_order","마스터오더","운송관리"),
    (24,"transport_schedule","운항스케줄","운송관리"),
    (25,"warehouse_receipt","입고","창고관리"),
    (26,"warehouse_release","출고","창고관리"),
    (27,"tracking_info","운송Tracking","Tracking"),
    (28,"customs_declaration","통관신고","Tracking"),
    (29,"invoice","청구서","회계/청구"),
    (30,"invoice_item","청구항목","회계/청구"),
    (31,"payment","입금","회계/청구"),
    (32,"transport_cost_master","운송원가마스터","회계/청구"),
    (33,"member_discount","회원별할인율","회계/청구"),
    (34,"tax_invoice","세금계산서","회계/청구"),
    (35,"notification","알림","시스템관리"),
    (36,"voc","VOC","시스템관리"),
    (37,"menu","메뉴","시스템관리"),
    (38,"role","권한","시스템관리"),
    (39,"menu_role","메뉴권한매핑","시스템관리"),
    (40,"individual_fare_rate","개인회원운임요율","기초정보관리"),
    (41,"corporate_fare_rate","법인회원운임요율","기초정보관리"),
    (42,"exchange_rate","환율","기초정보관리"),
    (43,"air_transport","항공운송수단","기초정보관리"),
    (44,"air_transport_cost","항공운송원가","기초정보관리"),
    (45,"sea_transport","해운운송수단","기초정보관리"),
    (46,"sea_transport_cost","해운운송원가","기초정보관리"),
    (47,"customs_broker","통관사","기초정보관리"),
    (48,"customs_broker_cost","통관원가","기초정보관리"),
    (49,"courier_transport_cost","택배사운송원가","기초정보관리"),
    (50,"courier_waybill","택배배송장","기초정보관리"),
]

COL_HEADERS = ["No","컬럼명(영문)","컬럼명(한글)","데이터타입","길이/정밀도","PK","FK","NOT NULL","기본값","설명"]
COL_WIDTHS  = [5,   22,          16,           14,          12,           6,   26,   10,        10,     40]

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_index_sheet(ws):
    ws.title = "목차"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # 타이틀
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "SNTL 통합 물류 플랫폼 — Data Dictionary  |  총 50개 테이블  |  PostgreSQL 16"
    t.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 헤더
    for ci, hdr in enumerate(["No","테이블명(영문)","테이블명(한글)","도메인"], 1):
        h_cell(ws, 2, ci, hdr)

    # 데이터
    for ri, row in enumerate(INDEX_DATA, 3):
        no, en, ko, domain = row
        bg = DOMAIN_COLORS.get(domain, "FFFFFF")
        d_cell(ws, ri, 1, no, bg=bg, align="center")
        d_cell(ws, ri, 2, en, bg=bg)
        d_cell(ws, ri, 3, ko, bg=bg)
        d_cell(ws, ri, 4, domain, bg=bg, align="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18


def write_domain_sheet(wb, domain_name, tables):
    safe_name = domain_name.replace("/", "_")
    ws = wb.create_sheet(title=safe_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    domain_bg = DOMAIN_COLORS.get(domain_name, "FFFFFF")

    current_row = 1
    for tbl in tables:
        en = tbl["name_en"]
        ko = tbl["name_ko"]
        desc = tbl["desc"]
        cols = tbl["cols"]

        # 테이블 제목
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(COL_HEADERS))
        title_cell = ws.cell(row=current_row, column=1,
                             value=f"{en}  ({ko})  —  {desc}")
        title_cell.font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
        title_cell.fill = PatternFill("solid", fgColor="1F3864")
        title_cell.alignment = Alignment(horizontal="left", vertical="center",
                                         indent=1)
        title_cell.border = border_thick()
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # 컬럼 헤더
        for ci, hdr in enumerate(COL_HEADERS, 1):
            h_cell(ws, current_row, ci, hdr, bg="2E75B6")
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # 데이터 행
        for row_data in cols:
            no, en_col, ko_col, dtype, prec, pk, fk, notnull, default, desc_col = row_data
            # 행 배경색 결정
            row_bg = domain_bg if no % 2 == 1 else ROW_ALT_BG
            if pk == "PK":
                row_bg = PK_COLOR
            elif fk and fk != "-":
                row_bg = FK_COLOR

            d_cell(ws, current_row, 1, no,       bg=row_bg, align="center")
            d_cell(ws, current_row, 2, en_col,    bg=row_bg)
            d_cell(ws, current_row, 3, ko_col,    bg=row_bg, align="center")
            d_cell(ws, current_row, 4, dtype,     bg=row_bg, align="center")
            d_cell(ws, current_row, 5, prec,      bg=row_bg, align="center")
            d_cell(ws, current_row, 6, pk,        bg=row_bg, align="center",
                   bold=(pk == "PK"))
            d_cell(ws, current_row, 7, fk,        bg=row_bg)
            d_cell(ws, current_row, 8, notnull,   bg=row_bg, align="center")
            d_cell(ws, current_row, 9, default,   bg=row_bg, align="center")
            d_cell(ws, current_row, 10, desc_col, bg=row_bg)
            ws.row_dimensions[current_row].height = 16
            current_row += 1

        current_row += 1  # 빈 행 간격

    set_col_widths(ws, COL_WIDTHS)


def write_enum_sheet(wb):
    ws = wb.create_sheet(title="Enum_요약")
    ws.sheet_view.showGridLines = False

    enums = [
        ("개인회원 등급",     "individual_member.grade",          "BASIC / SILVER / GOLD / VIP"),
        ("회원 상태",         "*.status",                          "ACTIVE / INACTIVE / WITHDRAWN"),
        ("법인 심사상태",     "corporate.approval_status",         "PENDING / APPROVED / REJECTED"),
        ("오더 상태",         "orders.order_status",               "REGISTERED / PACKED / WAREHOUSED / RELEASED / DELIVERED"),
        ("마스터오더 상태",   "master_order.master_status",        "CREATED / WAREHOUSED / RELEASED / IN_TRANSIT / DELIVERED"),
        ("서비스 구분",       "order_service.service_type",        "AIR / SEA / CIR / CCL"),
        ("청구서 상태",       "invoice.invoice_status",            "UNPAID / PARTIAL / PAID"),
        ("알림 발송상태",     "notification.send_status",          "PENDING / SENT / FAILED"),
        ("VOC 처리상태",      "voc.status",                        "OPEN / IN_PROGRESS / CLOSED"),
        ("선불금 충전상태",   "prepaid_charge_request.status",     "PENDING / CONFIRMED / REJECTED"),
        ("선불금 환불상태",   "prepaid_refund_request.status",     "PENDING / CONFIRMED / REJECTED"),
        ("운송원가 구분",     "*.cost_type",                       "VOLUME (부피기준) / WEIGHT (중량기준)"),
        ("통관사 API연동",    "customs_broker.api_integration",    "ON / OFF"),
        ("회원타입 구분",     "*.member_type",                     "INDIVIDUAL / CORPORATE"),
        ("오더번호 형식",     "orders.order_no",                   "ORD-YYYYMMDD-NNNNN"),
        ("마스터오더번호 형식","master_order.master_order_no",     "MO-YYYYMMDD-NNNNN"),
    ]

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "주요 Enum 값 요약"
    t.font = Font(name="맑은 고딕", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for ci, hdr in enumerate(["구분","컬럼","값"], 1):
        h_cell(ws, 2, ci, hdr)

    for ri, (cat, col, val) in enumerate(enums, 3):
        bg = "F9F9F9" if ri % 2 == 0 else "FFFFFF"
        d_cell(ws, ri, 1, cat, bg=bg, align="center")
        d_cell(ws, ri, 2, col, bg=bg)
        d_cell(ws, ri, 3, val, bg=bg)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 55


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 Sheet 제거

    # 1. 목차 시트
    ws_idx = wb.create_sheet(title="목차", index=0)
    write_index_sheet(ws_idx)

    # 2. 도메인별 시트
    domain_order = [
        "회원/인증", "코드관리", "오더관리", "운송관리",
        "창고관리", "Tracking", "회계/청구", "시스템관리", "기초정보관리"
    ]
    for domain in domain_order:
        if domain in TABLES:
            write_domain_sheet(wb, domain, TABLES[domain])

    # 3. Enum 요약 시트
    write_enum_sheet(wb)

    out_path = "SNTL_DataDictionary.xlsx"
    wb.save(out_path)
    print(f"[OK] {out_path} 생성 완료 (총 50개 테이블)")


if __name__ == "__main__":
    main()
