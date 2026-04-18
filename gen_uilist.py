import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'UI목록'

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_style(cell, bg='1E3A5F', fg='FFFFFF'):
    cell.font = Font(name='맑은 고딕', bold=True, color=fg, size=10)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

def data_style(cell, color='FFFFFF', center=False):
    cell.font = Font(name='맑은 고딕', size=9)
    cell.fill = PatternFill('solid', fgColor=color)
    cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
    cell.border = border

cat_colors = {
    '로그인/인증': 'D6EAF8',
    '회원관리': 'D5F5E3',
    '오더관리': 'FEF9E7',
    '마스터오더관리': 'FDEBD0',
    '창고관리': 'E8DAEF',
    '운송Tracking': 'D1F2EB',
    '회계/청구': 'FDFEFE',
    'VOC관리': 'EBF5FB',
    '고객지원': 'F9FBE7',
    '관리자': 'FDEDEC',
    '기초정보관리': 'DAE8FC',
}

# 타이틀
ws.merge_cells('A1:L1')
ws['A1'] = 'SNTL 통합 물류 플랫폼 — UI 화면 목록'
ws['A1'].font = Font(name='맑은 고딕', bold=True, size=14, color='FFFFFF')
ws['A1'].fill = PatternFill('solid', fgColor='1E3A5F')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

# 헤더 (row 2)
headers = ['화면ID', '화면명', '대분류', '중분류', '사용자유형', '주요기능', '관련 API', 'Phase', '연결화면', '비고', '우선순위', '상태']
ws.append(headers)
for i in range(1, 13):
    hdr_style(ws.cell(2, i), bg='2E86AB')
ws.row_dimensions[2].height = 22

screens = [
    ('SCR-001','로그인','로그인/인증','로그인','전체','ID/PW 입력, 소셜로그인, JWT Token 발급','/api/auth/login',1,'SCR-004, SCR-002','Access/Refresh Token 발급','P1','미개발'),
    ('SCR-002','ID 찾기','로그인/인증','ID/PW 찾기','회원','이름+휴대폰 입력, SMS 인증, ID 표시','/api/auth/find-id',1,'SCR-001','법인: 법인명+사업자번호','P1','미개발'),
    ('SCR-003','PW 재설정','로그인/인증','ID/PW 찾기','회원','SMS/이메일 인증, 새 비밀번호 설정','/api/auth/reset-password',1,'SCR-001','8자 이상 복잡성 조건','P1','미개발'),
    ('SCR-004','개인회원 가입','로그인/인증','회원가입','개인회원','본인확인, 개인정보 입력, 계정생성','/api/members/individual',1,'SCR-001','SMS 본인인증','P1','미개발'),
    ('SCR-005','법인회원 가입','로그인/인증','회원가입','법인관리자','법인정보 입력, 사업자등록증 첨부, 심사요청','/api/members/corporate',1,'SCR-001','첨부파일 업로드 포함','P1','미개발'),
    ('SCR-010','마이페이지','회원관리','개인회원','개인회원','회원정보 조회, 등급 확인, 수정 이동','/api/members/me',1,'SCR-011, SCR-012','','P1','미개발'),
    ('SCR-011','회원정보 수정','회원관리','개인회원','개인회원','이메일/주소/연락처 수정, PW 변경','/api/members/me',1,'SCR-010','','P1','미개발'),
    ('SCR-012','회원탈퇴','회원관리','회원탈퇴','회원','탈퇴 사유 선택, 본인확인, Soft Delete','/api/members/withdraw',1,'SCR-010','개인정보 Null 처리','P1','미개발'),
    ('SCR-013','법인 마이페이지','회원관리','법인회원','법인관리자','법인정보 조회, 부서/멤버 관리','/api/members/corporate/me',2,'SCR-014','','P2','미개발'),
    ('SCR-014','부서/멤버 관리','회원관리','법인회원','법인관리자','부서 CRUD, 멤버 초대/삭제','/api/corporate/departments',2,'SCR-013','','P2','미개발'),
    ('SCR-015','선불금 관리','회원관리','선불금 관리','개인회원/법인관리자','잔액 조회, 충전·환불 요청 이력, 버튼 이동','/api/prepaid/balance',1,'SCR-016, SCR-017','개인/법인 공통 화면','P1','미개발'),
    ('SCR-016','선불금 충전 요청','회원관리','선불금 관리','개인회원/법인관리자','충전 금액·입금자명·입금 예정일 입력, 요청 제출','/api/prepaid/charge',1,'SCR-015','입금 계좌 안내 표시','P1','미개발'),
    ('SCR-017','선불금 환불 요청','회원관리','선불금 관리','개인회원/법인관리자','환불 금액·계좌 정보 입력, 잔액 검증 후 요청','/api/prepaid/refund',1,'SCR-015','잔액 부족 시 요청 불가','P1','미개발'),
    ('SCR-020','오더 목록','오더관리','오더 기본관리','회원/운영자','오더 검색, 목록, 상태 필터, 페이징','/api/orders',1,'SCR-021, SCR-022','','P1','미개발'),
    ('SCR-021','오더 등록','오더관리','오더 기본관리','회원','송하인/수하인/화물정보 입력','/api/orders',1,'SCR-020, SCR-023','오더번호 자동생성','P1','미개발'),
    ('SCR-022','오더 상세/수정','오더관리','오더 기본관리','회원/운영자','오더 상세 조회, 수정, 삭제','/api/orders/{id}',1,'SCR-020','입고 후 수정불가','P1','미개발'),
    ('SCR-023','서비스 추가','오더관리','서비스 관리','회원','AIR/SEA/CIR/CCL 서비스 선택 및 정보 입력','/api/orders/{id}/services',1,'SCR-022','서비스별 입력항목 상이','P1','미개발'),
    ('SCR-024','운송비용 조회','오더관리','운송비용 관리','회원/운영자','자동계산 운송비, Extra Charge 입력','/api/orders/{id}/costs',2,'SCR-022','비용 계산식 적용','P2','미개발'),
    ('SCR-030','마스터오더 목록','마스터오더관리','마스터오더 조회','운영자','마스터오더 검색, 목록, 상태 필터','/api/master-orders',2,'SCR-031, SCR-032','','P2','미개발'),
    ('SCR-031','마스터오더 등록/패킹','마스터오더관리','마스터오더 등록','운영자','마스터오더 생성, 오더 패킹(추가/제거)','/api/master-orders',2,'SCR-030','','P2','미개발'),
    ('SCR-032','마스터오더 상세','마스터오더관리','마스터오더 조회','운영자','포함 오더 목록, 서비스 정보, Tracking 현황','/api/master-orders/{id}',2,'SCR-030','','P2','미개발'),
    ('SCR-040','창고 입고 처리','창고관리','입고 관리','운영자','바코드 스캔, 입고 확인 및 처리','/api/warehouse/receipt',2,'SCR-041','입고 후 오더 수정불가','P2','미개발'),
    ('SCR-041','창고 출고 처리','창고관리','출고 관리','운영자','바코드 스캔, 운송장 출력, 출고 처리','/api/warehouse/release',2,'SCR-040','운송장 PDF 출력','P2','미개발'),
    ('SCR-042','창고 현황','창고관리','입고 관리','운영자','입고 현황 조회, 재고 목록','/api/warehouse/status',2,'','','P2','미개발'),
    ('SCR-050','AIR Tracking','운송Tracking','항운(AIR) Tracking','회원/운영자','항공 화물 실시간 추적, 스케줄 조회','/api/tracking/air/{id}',2,'','외부 AIR API 연동','P2','미개발'),
    ('SCR-051','SEA Tracking','운송Tracking','해운(SEA) Tracking','회원/운영자','선박 운송 현황 추적, 선적 스케줄','/api/tracking/sea/{id}',2,'','외부 SEA API 연동','P2','미개발'),
    ('SCR-052','CIR Tracking','운송Tracking','택배(CIR) Tracking','회원/운영자','국제택배 실시간 추적','/api/tracking/cir/{id}',2,'','외부 택배사 API 연동','P2','미개발'),
    ('SCR-053','통관신고 현황','운송Tracking','통관신고 관리','회원/운영자','통관 신고 현황, 세관 응답 조회','/api/customs/{id}',2,'','CCL 서비스 연계','P2','미개발'),
    ('SCR-060','청구서 목록','회계/청구','청구서 관리','운영자/회원','청구서 목록, 검색, 상태 필터','/api/invoices',2,'SCR-061','','P2','미개발'),
    ('SCR-061','청구서 상세','회계/청구','청구서 관리','운영자/회원','청구 항목 상세, 입금 처리','/api/invoices/{id}',2,'SCR-060, SCR-062','','P2','미개발'),
    ('SCR-062','입금 처리','회계/청구','입금 관리','운영자','입금 확인, 미수금 처리','/api/invoices/{id}/payment',2,'SCR-061','','P2','미개발'),
    ('SCR-063','세금계산서','회계/청구','세금계산서','운영자','세금계산서 발행, 조회, 재발행','/api/tax-invoices',3,'SCR-060','','P3','미개발'),
    ('SCR-064','수입/비용 현황','회계/청구','수입 관리','관리자','기간별 수입, 원가, 수익 분석 차트','/api/statistics/revenue',3,'','차트 포함','P3','미개발'),
    ('SCR-065','운송원가 관리','회계/청구','운송원가 관리','관리자','서비스/구간별 원가 등록·수정','/api/transport-costs',2,'','','P2','미개발'),
    ('SCR-070','VOC 목록','VOC관리','VOC 조회','회원/운영자','VOC 목록, 상태 필터, 처리현황','/api/voc',2,'SCR-071, SCR-072','','P2','미개발'),
    ('SCR-071','VOC 등록','VOC관리','VOC 등록','회원','오더 선택, 불만/문의 내용 등록','/api/voc',2,'SCR-070','알림 자동 발송','P2','미개발'),
    ('SCR-072','VOC 답변','VOC관리','VOC 답변','운영자','VOC 내용 조회, 답변 등록, 상태 변경','/api/voc/{id}/reply',2,'SCR-070','고객 알림 발송','P2','미개발'),
    ('SCR-080','QnA 목록/등록','고객지원','QnA','회원','1:1 문의 목록, 문의 등록','/api/support/qna',2,'SCR-081','','P2','미개발'),
    ('SCR-081','QnA 답변','고객지원','QnA','운영자','관리자 답변 등록, 처리상태 변경','/api/support/qna/{id}',2,'SCR-080','','P2','미개발'),
    ('SCR-082','FAQ','고객지원','FAQ','전체','FAQ 카테고리별 조회, 관리자 CRUD','/api/support/faq',2,'','','P2','미개발'),
    ('SCR-083','공지사항','고객지원','공지사항','전체','공지사항 목록, 상세, 중요공지 상단고정','/api/support/notice',2,'','관리자: CRUD','P2','미개발'),
    ('SCR-090','관리자 대시보드','관리자','통계 관리','관리자','오더현황, 수익, 회원통계 요약 KPI+차트','/api/admin/dashboard',2,'','','P2','미개발'),
    ('SCR-091','회원 목록/관리','관리자','회원관리','관리자','회원 검색, 상세, 등급/상태 변경','/api/admin/members',1,'SCR-092','','P1','미개발'),
    ('SCR-092','법인 심사','관리자','회원관리','관리자','법인 가입 신청, 첨부파일 확인, 승인/거부','/api/admin/corporate/review',1,'SCR-091','승인 시 알림 발송','P1','미개발'),
    ('SCR-093','메뉴 관리','관리자','메뉴 관리','관리자','메뉴 트리 조회, 등록, 수정, 순서변경','/api/system/menus',2,'','계층형 트리','P2','미개발'),
    ('SCR-094','코드 관리','관리자','코드 관리','관리자','코드그룹 및 공통코드 CRUD','/api/system/codes',2,'','','P2','미개발'),
    ('SCR-095','권한 관리','관리자','권한 관리','관리자','역할 정의, 메뉴별 접근권한 설정','/api/system/roles',2,'','Role 기반 접근제어','P2','미개발'),
    ('SCR-096','택배사 관리','관리자','택배사 관리','관리자','택배사 CRUD, Tracking URL 관리','/api/system/couriers',2,'','','P2','미개발'),
    ('SCR-097','알림 관리','관리자','알림 관리','관리자','알림 발송 내역, 공지 알림 생성','/api/notifications',2,'','','P2','미개발'),
    ('SCR-098','통계','관리자','통계 관리','관리자','운송/비용 기간별 통계, 차트','/api/statistics',3,'','차트','P3','미개발'),
    ('SCR-099','데이터 백업','관리자','데이터 관리','관리자','DB 백업 실행, 백업 이력 조회','/api/admin/backup',3,'','','P3','미개발'),
    ('SCR-100','선불금 요청 관리','기초정보관리','선불금 관리','관리자','충전/환불 요청 목록 조회, PENDING 요청 확인·거부 처리','/api/prepaid/admin',1,'','확인 시 회원 잔액 자동 반영','P1','미개발'),
    ('SCR-101','운임요율 관리','기초정보관리','운임요율 관리','관리자','개인등급별(BASIC/SILVER/GOLD/VIP) 및 법인별 운임요율 CRUD, 적용기간 설정','/api/fare-rates',1,'','서비스구분(AIR/SEA/CIR/CCL) 탭','P1','미개발'),
    ('SCR-102','환율 관리','기초정보관리','환율 관리','관리자','최신 환율 목록 조회, 서울외국환중개소 API 수동 연계, 환율 수동 수정','/api/exchange-rates',1,'','영업일 자동 연계 이력 포함','P1','미개발'),
    ('SCR-103','항공운송수단 관리','기초정보관리','항공 운송수단','관리자','항공 운송구간(출발/도착 국가·공항) CRUD, 구간별 부피·중량기준 운송원가 CRUD','/api/air-transports',1,'','편명·원가 탭 분리','P1','미개발'),
    ('SCR-104','해운운송수단 관리','기초정보관리','해운 운송수단','관리자','해운 운송구간(출발/도착 국가·항구) CRUD, 구간별 부피·중량기준 운송원가 CRUD','/api/sea-transports',1,'','선사명·원가 탭 분리','P1','미개발'),
    ('SCR-105','통관사 관리','기초정보관리','통관사 관리','관리자','통관사 CRUD(국가·서비스구분·API연동ON/OFF), 통관원가 CRUD','/api/customs-brokers',1,'','API연동구분 ON/OFF 토글','P1','미개발'),
    ('SCR-106','택배배송장 관리','기초정보관리','택배배송장 관리','관리자','택배사별·국가별 배송장 양식 파일 업로드·다운로드·삭제','/api/courier-waybills',1,'SCR-096','MinIO 파일 저장','P1','미개발'),
]

for scr in screens:
    ws.append(list(scr))
    r = ws.max_row
    cat = scr[2]
    color = cat_colors.get(cat, 'FFFFFF')
    for c_idx in range(1, 13):
        cell = ws.cell(row=r, column=c_idx)
        center = c_idx in (1, 3, 4, 5, 8, 11, 12)
        data_style(cell, color=color, center=center)
    ws.row_dimensions[r].height = 20

col_widths = [12, 20, 14, 18, 16, 36, 30, 8, 22, 20, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 요약통계 시트
ws3 = wb.create_sheet('요약통계')
ws3.merge_cells('A1:D1')
ws3['A1'] = 'UI 화면 수 요약'
ws3['A1'].font = Font(name='맑은 고딕', bold=True, size=13, color='1E3A5F')
ws3.row_dimensions[1].height = 28

ws3.append(['대분류', '화면 수', '', ''])
for c in range(1,3):
    hdr_style(ws3.cell(2, c), bg='2E86AB')

cat_cnt = Counter(s[2] for s in screens)
for cat, cnt in cat_cnt.items():
    ws3.append([cat, cnt])
    r = ws3.max_row
    ws3.cell(r,1).fill = PatternFill('solid', fgColor=cat_colors.get(cat,'FFFFFF'))
    ws3.cell(r,1).font = Font(name='맑은 고딕', size=10)
    ws3.cell(r,2).font = Font(name='맑은 고딕', size=10)
    ws3.cell(r,2).alignment = Alignment(horizontal='center')

ws3.append(['합계', len(screens)])
r = ws3.max_row
ws3.cell(r,1).font = Font(name='맑은 고딕', bold=True)
ws3.cell(r,2).font = Font(name='맑은 고딕', bold=True)
ws3.cell(r,2).alignment = Alignment(horizontal='center')

ws3.append([])
ws3.append(['Phase', '화면 수'])
for c in range(1,3):
    hdr_style(ws3.cell(ws3.max_row, c), bg='2E86AB')

ph_cnt = Counter(s[7] for s in screens)
for ph in sorted(ph_cnt.keys()):
    ws3.append([f'Phase {ph}', ph_cnt[ph]])
    r = ws3.max_row
    ws3.cell(r,1).font = Font(name='맑은 고딕', size=10)
    ws3.cell(r,2).font = Font(name='맑은 고딕', size=10)
    ws3.cell(r,2).alignment = Alignment(horizontal='center')

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 12

wb.save('SNTL_UIList.xlsx')
print(f'Done. Total: {len(screens)} screens')
