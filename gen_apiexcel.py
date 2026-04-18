"""
SNTL API 목록 및 API 정의서 Excel 생성 스크립트
생성 파일:
  - SNTL_APIList.xlsx   : API 목록 (전체 테이블 + 도메인 통계)
  - SNTL_APISpec.xlsx   : API 정의서 (도메인별 시트)
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
from copy import copy

# ──────────────────────────────────────────────
# 색상 팔레트
# ──────────────────────────────────────────────
COLOR = {
    "header_dark":  "1F3864",  # 짙은 남색 (메인 헤더)
    "header_mid":   "2E75B6",  # 중간 청색 (섹션 헤더)
    "header_light": "9DC3E6",  # 연한 청색 (컬럼 헤더)
    "domain_auth":  "FFE699",  # 인증
    "domain_member":"C5E0B4",  # 회원
    "domain_order": "9DC3E6",  # 오더
    "domain_mo":    "F4B183",  # 마스터오더
    "domain_wh":    "D9E1F2",  # 창고
    "domain_tr":    "E2EFDA",  # Tracking
    "domain_bill":  "FCE4D6",  # 회계/청구
    "domain_cost":  "FFF2CC",  # 원가
    "domain_notif": "DDEBF7",  # 알림
    "domain_voc":   "F8CBAD",  # VOC
    "domain_sys":   "EDEDED",  # 시스템
    "domain_code":  "D6E4BC",  # 기준코드
    "method_get":   "70AD47",  # GET
    "method_post":  "4472C4",  # POST
    "method_put":   "ED7D31",  # PUT
    "method_patch": "7030A0",  # PATCH
    "method_delete":"FF0000",  # DELETE
    "phase1":       "C6EFCE",  # Phase 1
    "phase2":       "FFEB9C",  # Phase 2
    "phase3":       "FFC7CE",  # Phase 3
    "phase4":       "D9D9D9",  # Phase 4
    "white":        "FFFFFF",
    "light_gray":   "F2F2F2",
}

METHOD_COLOR = {
    "GET":    COLOR["method_get"],
    "POST":   COLOR["method_post"],
    "PUT":    COLOR["method_put"],
    "PATCH":  COLOR["method_patch"],
    "DELETE": COLOR["method_delete"],
}

PHASE_COLOR = {
    "1": COLOR["phase1"],
    "2": COLOR["phase2"],
    "3": COLOR["phase3"],
    "4": COLOR["phase4"],
}

DOMAIN_COLOR = {
    "인증":       COLOR["domain_auth"],
    "회원-개인":  COLOR["domain_member"],
    "회원-법인":  COLOR["domain_member"],
    "회원-담당자":COLOR["domain_member"],
    "회원-부서":  COLOR["domain_member"],
    "회원-등급":  COLOR["domain_member"],
    "선불금":     "E2EFDA",
    "오더":       COLOR["domain_order"],
    "마스터오더": COLOR["domain_mo"],
    "창고":       COLOR["domain_wh"],
    "Tracking":   COLOR["domain_tr"],
    "회계/청구":  COLOR["domain_bill"],
    "원가":       COLOR["domain_cost"],
    "알림":       COLOR["domain_notif"],
    "VOC":        COLOR["domain_voc"],
    "시스템":     COLOR["domain_sys"],
    "기준코드":   COLOR["domain_code"],
    "운임요율":   "FFF2CC",
    "환율":       "FFEB9C",
    "항공운송수단":"DAEEF3",
    "해운운송수단":"B8CCE4",
    "통관사":     "E4DFEC",
    "택배배송장":  "F2DCDB",
}

# ──────────────────────────────────────────────
# API 데이터 정의
# ──────────────────────────────────────────────
# 컬럼: [API_ID, 도메인, Method, URL, 기능명, 설명, 인증필요, 권한, Phase]
API_DATA = [
    # ── 인증
    ["API-A01","인증","POST","/auth/login","로그인","ID/PW로 로그인, Access+Refresh Token 발급","N","PUBLIC","1"],
    ["API-A02","인증","POST","/auth/logout","로그아웃","Redis Refresh Token 삭제, Access Token 블랙리스트","Y","AUTH","1"],
    ["API-A03","인증","POST","/auth/token/refresh","Token 갱신","Refresh Token으로 새 Access Token 발급","N","PUBLIC","1"],
    ["API-A04","인증","POST","/auth/sms/send","SMS 인증번호 발송","6자리 인증번호 발송 (유효 3분)","N","PUBLIC","1"],
    ["API-A05","인증","POST","/auth/sms/verify","SMS 인증번호 확인","인증번호 검증 후 verifyToken 발급","N","PUBLIC","1"],
    ["API-A06","인증","POST","/auth/password/reset-request","비밀번호 재설정 요청","본인 확인 후 resetToken 발급","N","PUBLIC","1"],
    ["API-A07","인증","PUT","/auth/password/reset","비밀번호 재설정","새 비밀번호로 변경 (8자+영+숫+특수)","N","PUBLIC","1"],
    # ── 회원-개인
    ["API-M01","회원-개인","POST","/members/individual","개인회원 가입","SMS 인증 완료 후 계정 생성","N","PUBLIC","1"],
    ["API-M02","회원-개인","GET","/members/individual","개인회원 목록","페이징 목록, 검색 지원","Y","ADMIN/OPR","1"],
    ["API-M03","회원-개인","GET","/members/individual/{memberId}","개인회원 상세","본인 또는 관리자","Y","SELF","1"],
    ["API-M04","회원-개인","PUT","/members/individual/{memberId}","개인회원 수정","loginId 변경 불가","Y","SELF","1"],
    ["API-M05","회원-개인","DELETE","/members/individual/{memberId}","개인회원 탈퇴","Soft Delete (개인정보 NULL)","Y","SELF","1"],
    # ── 회원-법인
    ["API-M06","회원-법인","POST","/members/corporate","법인회원 가입 신청","approval_status=PENDING, status=INACTIVE","N","PUBLIC","1"],
    ["API-M07","회원-법인","GET","/members/corporate","법인회원 목록","심사상태/상태 필터","Y","ADMIN/OPR","1"],
    ["API-M08","회원-법인","GET","/members/corporate/{corporateId}","법인회원 상세","본인 또는 관리자","Y","SELF/CORP","1"],
    ["API-M09","회원-법인","PUT","/members/corporate/{corporateId}","법인회원 수정","법인 정보 수정","Y","CORP_ADMIN","1"],
    ["API-M10","회원-법인","DELETE","/members/corporate/{corporateId}","법인회원 탈퇴","Soft Delete","Y","CORP_ADMIN","1"],
    ["API-M11","회원-법인","PATCH","/members/corporate/{corporateId}/approve","법인 심사 승인","승인+알림 발송, status→ACTIVE","Y","ADMIN/OPR","1"],
    ["API-M12","회원-법인","PATCH","/members/corporate/{corporateId}/reject","법인 심사 거부","거부 사유 포함 알림","Y","ADMIN/OPR","1"],
    ["API-M13","회원-법인","POST","/members/corporate/{corporateId}/attachments","첨부파일 업로드","MinIO 저장, PDF/JPG/PNG 10MB","Y","CORP/ADMIN","1"],
    ["API-M14","회원-법인","GET","/members/corporate/{corporateId}/attachments","첨부파일 목록","파일 목록 조회","Y","CORP/ADMIN","1"],
    ["API-M15","회원-법인","DELETE","/members/corporate/{corporateId}/attachments/{attachmentId}","첨부파일 삭제","MinIO 파일 삭제","Y","CORP_ADMIN","1"],
    # ── 회원-담당자
    ["API-M16","회원-담당자","POST","/members/corporate/{corporateId}/managers","담당자 추가","법인 담당자 추가","Y","CORP_ADMIN","1"],
    ["API-M17","회원-담당자","GET","/members/corporate/{corporateId}/managers","담당자 목록","담당자 목록 조회","Y","CORP/ADMIN","1"],
    ["API-M18","회원-담당자","PUT","/members/corporate/{corporateId}/managers/{managerId}","담당자 수정","담당자 정보 수정","Y","CORP_ADMIN","1"],
    ["API-M19","회원-담당자","DELETE","/members/corporate/{corporateId}/managers/{managerId}","담당자 삭제","담당자 삭제","Y","CORP_ADMIN","1"],
    # ── 회원-부서
    ["API-M20","회원-부서","POST","/members/corporate/{corporateId}/departments","부서 생성","법인 내 부서 생성","Y","CORP_ADMIN","1"],
    ["API-M21","회원-부서","GET","/members/corporate/{corporateId}/departments","부서 목록","부서 목록 조회","Y","CORP/ADMIN","1"],
    ["API-M22","회원-부서","PUT","/members/corporate/{corporateId}/departments/{deptId}","부서 수정","부서 정보 수정","Y","CORP_ADMIN","1"],
    ["API-M23","회원-부서","DELETE","/members/corporate/{corporateId}/departments/{deptId}","부서 삭제","부서 삭제","Y","CORP_ADMIN","1"],
    ["API-M24","회원-부서","POST","/members/corporate/{corporateId}/departments/{deptId}/managers","부서담당자 추가","부서 담당자 추가","Y","CORP_ADMIN","1"],
    ["API-M25","회원-부서","GET","/members/corporate/{corporateId}/departments/{deptId}/managers","부서담당자 목록","부서 담당자 목록","Y","CORP/ADMIN","1"],
    ["API-M26","회원-부서","PUT","/members/corporate/{corporateId}/departments/{deptId}/managers/{mgId}","부서담당자 수정","부서 담당자 수정","Y","CORP_ADMIN","1"],
    ["API-M27","회원-부서","DELETE","/members/corporate/{corporateId}/departments/{deptId}/managers/{mgId}","부서담당자 삭제","부서 담당자 삭제","Y","CORP_ADMIN","1"],
    # ── 회원-등급
    ["API-M28","회원-등급","GET","/members/grade-policies","등급 정책 목록","회원 등급별 정책 목록","Y","ADMIN","1"],
    ["API-M29","회원-등급","POST","/members/grade-policies","등급 정책 생성","새 등급 정책 등록","Y","ADMIN","1"],
    ["API-M30","회원-등급","PUT","/members/grade-policies/{policyId}","등급 정책 수정","등급 정책 수정","Y","ADMIN","1"],
    ["API-M31","회원-등급","DELETE","/members/grade-policies/{policyId}","등급 정책 삭제","등급 정책 삭제","Y","ADMIN","1"],
    # ── 오더
    ["API-O01","오더","POST","/orders","오더 등록","오더+송하인+수하인+화물 통합 등록","Y","AUTH","2"],
    ["API-O02","오더","GET","/orders","오더 목록","본인 오더 / 전체(관리자)","Y","AUTH","2"],
    ["API-O03","오더","GET","/orders/{orderId}","오더 상세","오더 전체 정보 조회","Y","AUTH","2"],
    ["API-O04","오더","PUT","/orders/{orderId}","오더 수정","WAREHOUSED 이후 불가","Y","SELF","2"],
    ["API-O05","오더","DELETE","/orders/{orderId}","오더 삭제","WAREHOUSED 이후 불가","Y","SELF","2"],
    ["API-O06","오더","GET","/orders/{orderId}/shipper","송하인 정보 조회","송하인 상세 조회","Y","AUTH","2"],
    ["API-O07","오더","PUT","/orders/{orderId}/shipper","송하인 정보 수정","송하인 정보 수정","Y","SELF","2"],
    ["API-O08","오더","GET","/orders/{orderId}/consignee","수하인 정보 조회","수하인 상세 조회","Y","AUTH","2"],
    ["API-O09","오더","PUT","/orders/{orderId}/consignee","수하인 정보 수정","수하인 정보 수정","Y","SELF","2"],
    ["API-O10","오더","GET","/orders/{orderId}/cargo","화물 정보 조회","화물 상세 조회","Y","AUTH","2"],
    ["API-O11","오더","PUT","/orders/{orderId}/cargo","화물 정보 수정","화물 정보 수정","Y","SELF","2"],
    ["API-O12","오더","POST","/orders/{orderId}/services","서비스 추가","AIR/SEA/CIR/CCL 서비스 추가","Y","SELF","2"],
    ["API-O13","오더","GET","/orders/{orderId}/services","서비스 목록","오더 서비스 목록","Y","AUTH","2"],
    ["API-O14","오더","PUT","/orders/{orderId}/services/{serviceId}","서비스 수정","서비스 정보 수정","Y","SELF","2"],
    ["API-O15","오더","DELETE","/orders/{orderId}/services/{serviceId}","서비스 삭제","서비스 삭제","Y","SELF","2"],
    # ── 마스터오더
    ["API-MO01","마스터오더","POST","/master-orders","마스터오더 생성","오더 패킹, MO번호 자동 생성","Y","ADMIN/OPR","2"],
    ["API-MO02","마스터오더","GET","/master-orders","마스터오더 목록","페이징+필터","Y","ADMIN/OPR","2"],
    ["API-MO03","마스터오더","GET","/master-orders/{masterOrderId}","마스터오더 상세","상세 조회","Y","ADMIN/OPR","2"],
    ["API-MO04","마스터오더","PUT","/master-orders/{masterOrderId}","마스터오더 수정","마스터오더 수정","Y","ADMIN/OPR","2"],
    ["API-MO05","마스터오더","DELETE","/master-orders/{masterOrderId}","마스터오더 삭제","마스터오더 삭제","Y","ADMIN","2"],
    ["API-MO06","마스터오더","GET","/master-orders/{masterOrderId}/schedules","운항 스케줄 조회","운항 스케줄 조회","Y","ADMIN/OPR","2"],
    ["API-MO07","마스터오더","POST","/master-orders/{masterOrderId}/schedules","운항 스케줄 등록","운항 스케줄 등록","Y","ADMIN/OPR","2"],
    ["API-MO08","마스터오더","PUT","/master-orders/{masterOrderId}/schedules/{scheduleId}","운항 스케줄 수정","운항 스케줄 수정","Y","ADMIN/OPR","2"],
    # ── 창고
    ["API-W01","창고","POST","/warehouse/receipts","입고 등록","바코드 스캔, 오더 상태→WAREHOUSED","Y","ADMIN/OPR","2"],
    ["API-W02","창고","GET","/warehouse/receipts","입고 목록","입고 이력 목록","Y","ADMIN/OPR","2"],
    ["API-W03","창고","GET","/warehouse/receipts/{receiptId}","입고 상세","입고 상세 조회","Y","ADMIN/OPR","2"],
    ["API-W04","창고","POST","/warehouse/releases","출고 등록","마스터오더 단위 출고 처리","Y","ADMIN/OPR","2"],
    ["API-W05","창고","GET","/warehouse/releases","출고 목록","출고 이력 목록","Y","ADMIN/OPR","2"],
    ["API-W06","창고","GET","/warehouse/releases/{releaseId}","출고 상세","출고 상세 조회","Y","ADMIN/OPR","2"],
    ["API-W07","창고","GET","/warehouse/barcode/{barcode}","바코드 스캔 조회","바코드로 오더 즉시 조회","Y","ADMIN/OPR","2"],
    # ── Tracking
    ["API-T01","Tracking","POST","/tracking","트래킹 등록","운송 상태 이력 등록 (갱신 임계 30분)","Y","ADMIN/OPR","3"],
    ["API-T02","Tracking","GET","/tracking/{masterOrderId}","트래킹 이력 조회","전체 이력 최신순 조회","Y","AUTH","3"],
    ["API-T03","Tracking","PUT","/tracking/{trackingId}","트래킹 상태 수정","트래킹 정보 수정","Y","ADMIN/OPR","3"],
    ["API-T04","Tracking","POST","/tracking/customs","통관신고 등록","통관신고 정보 등록","Y","ADMIN/OPR","3"],
    ["API-T05","Tracking","GET","/tracking/customs/{masterOrderId}","통관신고 조회","통관신고 정보 조회","Y","AUTH","3"],
    ["API-T06","Tracking","PUT","/tracking/customs/{declarationId}","통관신고 수정","통관신고 수정","Y","ADMIN/OPR","3"],
    # ── 회계/청구
    ["API-B01","회계/청구","POST","/billing/invoices","청구서 생성","운송비 자동 계산, INV번호 생성","Y","ADMIN/OPR","3"],
    ["API-B02","회계/청구","GET","/billing/invoices","청구서 목록","청구서 목록","Y","AUTH","3"],
    ["API-B03","회계/청구","GET","/billing/invoices/{invoiceId}","청구서 상세","청구서 상세 조회","Y","AUTH","3"],
    ["API-B04","회계/청구","PUT","/billing/invoices/{invoiceId}","청구서 수정","확정 전 수정 가능","Y","ADMIN/OPR","3"],
    ["API-B05","회계/청구","PATCH","/billing/invoices/{invoiceId}/confirm","청구서 확정","확정 후 수정 불가","Y","ADMIN/OPR","3"],
    ["API-B06","회계/청구","GET","/billing/invoices/{invoiceId}/items","청구 항목 목록","항목별 목록 조회","Y","AUTH","3"],
    ["API-B07","회계/청구","POST","/billing/invoices/{invoiceId}/items","청구 항목 추가","청구 항목 추가","Y","ADMIN/OPR","3"],
    ["API-B08","회계/청구","PUT","/billing/invoices/{invoiceId}/items/{itemId}","청구 항목 수정","항목 수정","Y","ADMIN/OPR","3"],
    ["API-B09","회계/청구","DELETE","/billing/invoices/{invoiceId}/items/{itemId}","청구 항목 삭제","항목 삭제","Y","ADMIN/OPR","3"],
    ["API-B10","회계/청구","POST","/billing/payments","입금 등록","전액/일부 입금, status 자동 갱신","Y","ADMIN/OPR","3"],
    ["API-B11","회계/청구","GET","/billing/payments","입금 목록","입금 이력 목록","Y","ADMIN/OPR","3"],
    ["API-B12","회계/청구","GET","/billing/payments/{paymentId}","입금 상세","입금 상세 조회","Y","ADMIN/OPR","3"],
    # ── 원가
    ["API-C01","원가","GET","/costs/transport-masters","운송원가 목록","서비스/구간별 원가 목록","Y","ADMIN/OPR","3"],
    ["API-C02","원가","POST","/costs/transport-masters","운송원가 등록","운송원가 등록","Y","ADMIN","3"],
    ["API-C03","원가","GET","/costs/transport-masters/{costId}","운송원가 상세","운송원가 상세 조회","Y","ADMIN/OPR","3"],
    ["API-C04","원가","PUT","/costs/transport-masters/{costId}","운송원가 수정","운송원가 수정","Y","ADMIN","3"],
    ["API-C05","원가","DELETE","/costs/transport-masters/{costId}","운송원가 삭제","운송원가 삭제","Y","ADMIN","3"],
    ["API-C06","원가","GET","/costs/discounts","할인율 목록","회원별 할인율 목록","Y","ADMIN/OPR","3"],
    ["API-C07","원가","POST","/costs/discounts","할인율 등록","회원별 할인율 등록","Y","ADMIN","3"],
    ["API-C08","원가","PUT","/costs/discounts/{discountId}","할인율 수정","할인율 수정","Y","ADMIN","3"],
    ["API-C09","원가","DELETE","/costs/discounts/{discountId}","할인율 삭제","할인율 삭제","Y","ADMIN","3"],
    ["API-C10","원가","POST","/costs/calculate","운송비 계산","운송비 미리보기 (청구 전)","Y","AUTH","3"],
    # ── 알림
    ["API-N01","알림","GET","/notifications","알림 목록","본인 알림 최신순 조회","Y","AUTH","4"],
    ["API-N02","알림","PATCH","/notifications/{notificationId}/read","알림 읽음 처리","단건 읽음 처리","Y","SELF","4"],
    ["API-N03","알림","PATCH","/notifications/read-all","전체 알림 읽음","전체 읽음 처리","Y","AUTH","4"],
    ["API-N04","알림","DELETE","/notifications/{notificationId}","알림 삭제","알림 삭제","Y","SELF","4"],
    # ── VOC
    ["API-V01","VOC","POST","/voc","VOC 등록","민원/불만 등록","Y","AUTH","3"],
    ["API-V02","VOC","GET","/voc","VOC 목록","목록 조회 (본인/전체)","Y","AUTH","3"],
    ["API-V03","VOC","GET","/voc/{vocId}","VOC 상세","상세 조회","Y","AUTH","3"],
    ["API-V04","VOC","PUT","/voc/{vocId}","VOC 수정","OPEN 상태에서만 수정","Y","SELF","3"],
    ["API-V05","VOC","PATCH","/voc/{vocId}/status","VOC 상태 변경","OPEN→IN_PROGRESS→CLOSED","Y","ADMIN/OPR","3"],
    ["API-V06","VOC","DELETE","/voc/{vocId}","VOC 삭제","VOC 삭제","Y","SELF","3"],
    # ── 시스템
    ["API-S01","시스템","GET","/system/menus","메뉴 목록","전체 메뉴 Tree 조회","Y","ADMIN","4"],
    ["API-S02","시스템","POST","/system/menus","메뉴 등록","메뉴 등록","Y","ADMIN","4"],
    ["API-S03","시스템","PUT","/system/menus/{menuId}","메뉴 수정","메뉴 수정","Y","ADMIN","4"],
    ["API-S04","시스템","DELETE","/system/menus/{menuId}","메뉴 삭제","메뉴 삭제","Y","ADMIN","4"],
    ["API-S05","시스템","GET","/system/roles","역할 목록","권한 역할 목록","Y","ADMIN","4"],
    ["API-S06","시스템","POST","/system/roles","역할 등록","권한 역할 등록","Y","ADMIN","4"],
    ["API-S07","시스템","PUT","/system/roles/{roleId}","역할 수정","권한 역할 수정","Y","ADMIN","4"],
    ["API-S08","시스템","DELETE","/system/roles/{roleId}","역할 삭제","권한 역할 삭제","Y","ADMIN","4"],
    ["API-S09","시스템","POST","/system/menu-roles","메뉴-역할 매핑 등록","메뉴-권한 매핑","Y","ADMIN","4"],
    ["API-S10","시스템","DELETE","/system/menu-roles/{menuRoleId}","메뉴-역할 매핑 삭제","매핑 삭제","Y","ADMIN","4"],
    ["API-S11","시스템","GET","/system/code-groups","코드그룹 목록","코드그룹 목록","Y","ADMIN/OPR","4"],
    ["API-S12","시스템","POST","/system/code-groups","코드그룹 등록","코드그룹 등록","Y","ADMIN","4"],
    ["API-S13","시스템","PUT","/system/code-groups/{groupId}","코드그룹 수정","코드그룹 수정","Y","ADMIN","4"],
    ["API-S14","시스템","GET","/system/codes","공통코드 목록","그룹별 공통코드 목록","Y","AUTH","4"],
    ["API-S15","시스템","POST","/system/codes","공통코드 등록","공통코드 등록","Y","ADMIN","4"],
    ["API-S16","시스템","PUT","/system/codes/{codeId}","공통코드 수정","공통코드 수정","Y","ADMIN","4"],
    ["API-S17","시스템","DELETE","/system/codes/{codeId}","공통코드 삭제","공통코드 삭제","Y","ADMIN","4"],
    # ── 기준코드
    ["API-R01","기준코드","GET","/codes/countries","국가코드 목록","전체 국가코드 목록","Y","AUTH","1"],
    ["API-R02","기준코드","POST","/codes/countries","국가코드 등록","국가코드 등록","Y","ADMIN","1"],
    ["API-R03","기준코드","PUT","/codes/countries/{id}","국가코드 수정","국가코드 수정","Y","ADMIN","1"],
    ["API-R04","기준코드","DELETE","/codes/countries/{id}","국가코드 삭제","국가코드 삭제","Y","ADMIN","1"],
    ["API-R05","기준코드","GET","/codes/airports","공항코드 목록","전체 공항코드 목록","Y","AUTH","1"],
    ["API-R06","기준코드","POST","/codes/airports","공항코드 등록","공항코드 등록","Y","ADMIN","1"],
    ["API-R07","기준코드","PUT","/codes/airports/{id}","공항코드 수정","공항코드 수정","Y","ADMIN","1"],
    ["API-R08","기준코드","DELETE","/codes/airports/{id}","공항코드 삭제","공항코드 삭제","Y","ADMIN","1"],
    ["API-R09","기준코드","GET","/codes/ports","항구코드 목록","전체 항구코드 목록","Y","AUTH","1"],
    ["API-R10","기준코드","POST","/codes/ports","항구코드 등록","항구코드 등록","Y","ADMIN","1"],
    ["API-R11","기준코드","PUT","/codes/ports/{id}","항구코드 수정","항구코드 수정","Y","ADMIN","1"],
    ["API-R12","기준코드","DELETE","/codes/ports/{id}","항구코드 삭제","항구코드 삭제","Y","ADMIN","1"],
    ["API-R13","기준코드","GET","/codes/airlines","항공사코드 목록","전체 항공사코드 목록","Y","AUTH","1"],
    ["API-R14","기준코드","POST","/codes/airlines","항공사코드 등록","항공사코드 등록","Y","ADMIN","1"],
    ["API-R15","기준코드","PUT","/codes/airlines/{id}","항공사코드 수정","항공사코드 수정","Y","ADMIN","1"],
    ["API-R16","기준코드","DELETE","/codes/airlines/{id}","항공사코드 삭제","항공사코드 삭제","Y","ADMIN","1"],
    ["API-R17","기준코드","GET","/codes/couriers","택배사 목록","전체 택배사 목록","Y","AUTH","1"],
    ["API-R18","기준코드","POST","/codes/couriers","택배사 등록","택배사 등록","Y","ADMIN","1"],
    ["API-R19","기준코드","PUT","/codes/couriers/{id}","택배사 수정","택배사 수정","Y","ADMIN","1"],
    ["API-R20","기준코드","DELETE","/codes/couriers/{id}","택배사 삭제","택배사 삭제","Y","ADMIN","1"],
    # ── 운임요율
    ["API-FR01","운임요율","GET","/fare-rates/individual","개인회원 등급별 운임요율 목록","회원등급(BASIC/SILVER/GOLD/VIP)별 운임요율(%) 목록","Y","ADMIN/OPR","1"],
    ["API-FR02","운임요율","POST","/fare-rates/individual","개인회원 운임요율 등록","등급별 운임요율 등록","Y","ADMIN","1"],
    ["API-FR03","운임요율","PUT","/fare-rates/individual/{rateId}","개인회원 운임요율 수정","운임요율 수정","Y","ADMIN","1"],
    ["API-FR04","운임요율","DELETE","/fare-rates/individual/{rateId}","개인회원 운임요율 삭제","운임요율 삭제","Y","ADMIN","1"],
    ["API-FR05","운임요율","GET","/fare-rates/corporate","법인회원별 운임요율 목록","법인ID별 운임요율 목록","Y","ADMIN/OPR","1"],
    ["API-FR06","운임요율","POST","/fare-rates/corporate","법인회원 운임요율 등록","법인별 운임요율 등록","Y","ADMIN","1"],
    ["API-FR07","운임요율","PUT","/fare-rates/corporate/{rateId}","법인회원 운임요율 수정","운임요율 수정","Y","ADMIN","1"],
    ["API-FR08","운임요율","DELETE","/fare-rates/corporate/{rateId}","법인회원 운임요율 삭제","운임요율 삭제","Y","ADMIN","1"],
    # ── 환율
    ["API-EX01","환율","GET","/exchange-rates","환율 조회","서울외국환중개소 기준환율. 로그인 국가코드 기준","Y","AUTH","1"],
    ["API-EX02","환율","POST","/exchange-rates/sync","환율 연계","서울외국환중개소 API 호출, 영업일 오전 8시 자동 갱신","Y","ADMIN","1"],
    # ── 항공운송수단
    ["API-AT01","항공운송수단","GET","/transports/air","항공운송수단 목록","출발/도착 공항코드로 조회","Y","ADMIN/OPR","1"],
    ["API-AT02","항공운송수단","POST","/transports/air","항공운송수단 등록","출발/도착 공항코드+항공편명 등록","Y","ADMIN","1"],
    ["API-AT03","항공운송수단","PUT","/transports/air/{transportId}","항공운송수단 수정","항공편명 수정","Y","ADMIN","1"],
    ["API-AT04","항공운송수단","DELETE","/transports/air/{transportId}","항공운송수단 삭제","운송원가 연결 시 삭제 불가","Y","ADMIN","1"],
    ["API-AT05","항공운송수단","GET","/transports/air/{transportId}/costs","항공 운송원가 조회","부피/중량기준 운송원가 조회","Y","ADMIN/OPR","1"],
    ["API-AT06","항공운송수단","POST","/transports/air/{transportId}/costs","항공 운송원가 등록","부피기준(X·Y·Z·USD) 또는 중량기준(Kg·USD)","Y","ADMIN","1"],
    ["API-AT07","항공운송수단","PUT","/transports/air/{transportId}/costs/{costId}","항공 운송원가 수정","운송원가(USD) 수정","Y","ADMIN","1"],
    # ── 해운운송수단
    ["API-ST01","해운운송수단","GET","/transports/sea","해운운송수단 목록","출발/도착 항구코드로 조회","Y","ADMIN/OPR","1"],
    ["API-ST02","해운운송수단","POST","/transports/sea","해운운송수단 등록","출발/도착 항구코드+해운편명 등록","Y","ADMIN","1"],
    ["API-ST03","해운운송수단","PUT","/transports/sea/{transportId}","해운운송수단 수정","해운편명 수정","Y","ADMIN","1"],
    ["API-ST04","해운운송수단","DELETE","/transports/sea/{transportId}","해운운송수단 삭제","운송원가 연결 시 삭제 불가","Y","ADMIN","1"],
    ["API-ST05","해운운송수단","GET","/transports/sea/{transportId}/costs","해운 운송원가 조회","부피/중량기준 운송원가 조회","Y","ADMIN/OPR","1"],
    ["API-ST06","해운운송수단","POST","/transports/sea/{transportId}/costs","해운 운송원가 등록","부피기준(X·Y·Z·USD) 또는 중량기준(Kg·USD)","Y","ADMIN","1"],
    ["API-ST07","해운운송수단","PUT","/transports/sea/{transportId}/costs/{costId}","해운 운송원가 수정","운송원가(USD) 수정","Y","ADMIN","1"],
    # ── 통관사
    ["API-CB01","통관사","GET","/customs-brokers","통관사 목록","국가코드·운송구분·입항코드로 조회","Y","ADMIN/OPR","1"],
    ["API-CB02","통관사","POST","/customs-brokers","통관사 등록","국가코드+서비스구분(AIR/SEA)+입항코드+API연동여부","Y","ADMIN","1"],
    ["API-CB03","통관사","PUT","/customs-brokers/{brokerId}","통관사 수정","통관사 정보 수정","Y","ADMIN","1"],
    ["API-CB04","통관사","DELETE","/customs-brokers/{brokerId}","통관사 삭제","통관원가 연결 시 삭제 불가","Y","ADMIN","1"],
    ["API-CB05","통관사","GET","/customs-brokers/{brokerId}/costs","통관원가 조회","부피/중량기준 통관원가 조회","Y","ADMIN/OPR","1"],
    ["API-CB06","통관사","POST","/customs-brokers/{brokerId}/costs","통관원가 등록","부피기준(X·Y·Z·USD) 또는 중량기준(Kg·USD)","Y","ADMIN","1"],
    ["API-CB07","통관사","PUT","/customs-brokers/{brokerId}/costs/{costId}","통관원가 수정","통관원가(USD) 수정","Y","ADMIN","1"],
    # ── 택배배송장
    ["API-CW01","택배배송장","GET","/codes/couriers/{courierId}/waybills","택배배송장 목록","국가코드·택배사명으로 조회","Y","ADMIN/OPR","1"],
    ["API-CW02","택배배송장","POST","/codes/couriers/{courierId}/waybills","택배배송장 등록","국가코드+배송장 양식파일(BLOB) 등록","Y","ADMIN","1"],
    ["API-CW03","택배배송장","PUT","/codes/couriers/{courierId}/waybills/{waybillId}","택배배송장 수정","배송장 양식파일 수정","Y","ADMIN","1"],
    ["API-CW04","택배배송장","DELETE","/codes/couriers/{courierId}/waybills/{waybillId}","택배배송장 삭제","배송장 삭제","Y","ADMIN","1"],
    # ── 선불금
    ["API-PP01","선불금","GET","/prepaid/balance","선불금 잔액 조회 (본인)","본인 선불금 잔액 조회","Y","AUTH","1"],
    ["API-PP02","선불금","POST","/prepaid/charge-request","선불금 충전 요청","입금 후 충전 금액 입력, 관리자 확인 요청","Y","AUTH","1"],
    ["API-PP03","선불금","PATCH","/prepaid/charge-requests/{requestId}/confirm","충전 확인 (관리자)","실제 입금 확인 후 최종 충전 처리, 잔액 반영","Y","ADMIN/OPR","1"],
    ["API-PP04","선불금","POST","/prepaid/refund-request","환불 요청","미사용 잔액 환불 요청, 환불 계좌정보 입력","Y","AUTH","1"],
    ["API-PP05","선불금","GET","/prepaid/admin/balances","전체 선불금 조회 (관리자)","회원별·법인별 선불금 잔액 전체 조회","Y","ADMIN/OPR","1"],
]

# ──────────────────────────────────────────────
# API 정의서 상세 데이터
# 컬럼: [API_ID, 기능명, Method, URL, 설명, Path Params, Query Params, Request Body 요약, Response 200 요약, 주요 에러코드]
# ──────────────────────────────────────────────
API_SPEC_DATA = [
    ["API-A01","로그인","POST","/auth/login",
     "ID/PW 로그인. Access Token(30분) + Refresh Token(7일) 발급. 5회 실패 시 30분 계정 잠금",
     "-",
     "-",
     "loginId(String,필수), password(String,필수)",
     "accessToken, refreshToken, tokenType, expiresIn, memberType, memberId, loginId, name, role",
     "PASSWORD_MISMATCH(400), MEMBER_NOT_FOUND(404), ACCOUNT_LOCKED(423)"],
    ["API-A02","로그아웃","POST","/auth/logout",
     "Redis Refresh Token 삭제 + Access Token 블랙리스트 등록",
     "-","-","-","success: true","TOKEN_INVALID(401)"],
    ["API-A03","Token 갱신","POST","/auth/token/refresh",
     "Refresh Token으로 새 Access Token 발급",
     "-","-",
     "refreshToken(String,필수)",
     "accessToken, expiresIn",
     "TOKEN_EXPIRED(401), TOKEN_INVALID(401)"],
    ["API-A04","SMS 인증번호 발송","POST","/auth/sms/send",
     "6자리 인증번호 SMS 발송. 유효시간 3분",
     "-","-",
     "phone(String,필수), purpose(SIGNUP|PASSWORD_RESET|ID_FIND,필수)",
     "sessionKey",
     "INVALID_INPUT(400)"],
    ["API-A05","SMS 인증번호 확인","POST","/auth/sms/verify",
     "인증번호 검증 후 verifyToken 발급",
     "-","-",
     "sessionKey(String,필수), code(String,필수)",
     "verified(boolean), verifyToken",
     "SMS_CODE_EXPIRED(400), SMS_CODE_INVALID(400)"],
    ["API-A06","비밀번호 재설정 요청","POST","/auth/password/reset-request",
     "본인 확인 후 resetToken 발급",
     "-","-",
     "loginId(String,필수), verifyToken(String,필수)",
     "resetToken",
     "MEMBER_NOT_FOUND(404)"],
    ["API-A07","비밀번호 재설정","PUT","/auth/password/reset",
     "새 비밀번호로 변경 (영문+숫자+특수문자 8자 이상)",
     "-","-",
     "resetToken(필수), newPassword(필수), newPasswordConfirm(필수)",
     "success: true",
     "INVALID_INPUT(400)"],
    ["API-M01","개인회원 가입","POST","/members/individual",
     "SMS 인증 완료 후 계정 생성",
     "-","-",
     "loginId(필수), password(필수), name, birthDate, phone, email, zipcode, address, addressDetail, verifyToken(필수)",
     "memberId, loginId, name, grade(BASIC), status(ACTIVE), createdAt",
     "DUPLICATE_LOGIN_ID(409), SMS_CODE_EXPIRED(400)"],
    ["API-M02","개인회원 목록","GET","/members/individual",
     "페이징 목록. 이름/아이디/이메일 검색",
     "-",
     "page, size, sort, keyword, status",
     "-",
     "content[], totalElements, totalPages, size, number",
     "FORBIDDEN(403)"],
    ["API-M03","개인회원 상세","GET","/members/individual/{memberId}",
     "본인 또는 관리자만 조회",
     "memberId(Long,필수)","-","-",
     "memberId, loginId, name, birthDate, phone, email, grade, balance, status, createdAt",
     "MEMBER_NOT_FOUND(404), FORBIDDEN(403)"],
    ["API-M04","개인회원 수정","PUT","/members/individual/{memberId}",
     "loginId 변경 불가. 비밀번호 변경 시 currentPassword 필수",
     "memberId(Long,필수)","-",
     "name, birthDate, phone, email, zipcode, address, addressDetail, currentPassword, newPassword",
     "memberId, updatedAt",
     "MEMBER_NOT_FOUND(404), PASSWORD_MISMATCH(400)"],
    ["API-M05","개인회원 탈퇴","DELETE","/members/individual/{memberId}",
     "Soft Delete. 개인정보 NULL 처리, id/login_id 보존, status→WITHDRAWN",
     "memberId(Long,필수)","-",
     "password(필수), reason",
     "success: true",
     "PASSWORD_MISMATCH(400), MEMBER_NOT_FOUND(404)"],
    ["API-M06","법인회원 가입 신청","POST","/members/corporate",
     "초기 approval_status=PENDING, status=INACTIVE. 관리자 심사 필요",
     "-","-",
     "loginId(필수), password(필수), corpName(필수), ceoName, businessNumber, phone, email, zipcode, address, addressDetail",
     "corporateId, loginId, corpName, approvalStatus(PENDING), status(INACTIVE)",
     "DUPLICATE_LOGIN_ID(409)"],
    ["API-M11","법인 심사 승인","PATCH","/members/corporate/{corporateId}/approve",
     "approval_status→APPROVED, status→ACTIVE. 승인 알림 발송",
     "corporateId(Long,필수)","-",
     "memo",
     "corporateId, approvalStatus(APPROVED), status(ACTIVE)",
     "MEMBER_NOT_FOUND(404), FORBIDDEN(403)"],
    ["API-M12","법인 심사 거부","PATCH","/members/corporate/{corporateId}/reject",
     "approval_status→REJECTED. 거부 사유 포함 알림 발송",
     "corporateId(Long,필수)","-",
     "rejectReason(필수)",
     "corporateId, approvalStatus(REJECTED)",
     "MEMBER_NOT_FOUND(404), FORBIDDEN(403)"],
    ["API-M13","첨부파일 업로드","POST","/members/corporate/{corporateId}/attachments",
     "MinIO 저장. 최대 10MB. PDF/JPG/PNG 허용",
     "corporateId(Long,필수)","-",
     "[multipart] file(필수), fileType(BUSINESS_LICENSE|OTHER,필수)",
     "attachmentId, fileName, fileType, fileSize, fileUrl, uploadedAt",
     "INVALID_INPUT(400)"],
    ["API-O01","오더 등록","POST","/orders",
     "오더+송하인+수하인+화물 통합 등록. ORD-YYYYMMDD-NNNNN 자동 생성",
     "-","-",
     "shipper{name,phone,email,country,zipcode,address,addressDetail}, consignee{name,phone,email,country,zipcode,address,...}, cargo{cargoName,hsCode,quantity,totalWeight,volume*,declaredValue,currency}, remark",
     "orderId, orderNo, orderStatus(REGISTERED), createdAt",
     "INVALID_INPUT(400), CORPORATE_NOT_APPROVED(403)"],
    ["API-O02","오더 목록","GET","/orders",
     "일반 사용자: 본인 오더만. ADMIN/OPR: 전체",
     "-",
     "page, size, orderStatus, startDate, endDate, keyword",
     "-",
     "content[], totalElements, totalPages",
     "FORBIDDEN(403)"],
    ["API-O03","오더 상세","GET","/orders/{orderId}",
     "오더 전체 정보 (shipper, consignee, cargo, services 포함)",
     "orderId(Long,필수)","-","-",
     "orderId, orderNo, orderStatus, shipper{}, consignee{}, cargo{cbm계산포함}, services[]",
     "ORDER_NOT_FOUND(404)"],
    ["API-O04","오더 수정","PUT","/orders/{orderId}",
     "WAREHOUSED 이후 수정 불가",
     "orderId(Long,필수)","-",
     "오더 등록과 동일 구조",
     "orderId, updatedAt",
     "ORDER_ALREADY_WAREHOUSED(409), FORBIDDEN(403)"],
    ["API-O05","오더 삭제","DELETE","/orders/{orderId}",
     "WAREHOUSED 이후 삭제 불가",
     "orderId(Long,필수)","-","-","success: true",
     "ORDER_ALREADY_WAREHOUSED(409), FORBIDDEN(403)"],
    ["API-O12","서비스 추가","POST","/orders/{orderId}/services",
     "AIR/SEA/CIR/CCL 유형별 세부정보 입력",
     "orderId(Long,필수)","-",
     "serviceType(AIR|SEA|CIR|CCL,필수) + 유형별 세부필드 (flightNo, vesselName, trackingNo, declarationNo 등)",
     "serviceId, orderId, serviceType, createdAt",
     "ORDER_NOT_FOUND(404), ORDER_ALREADY_WAREHOUSED(409)"],
    ["API-MO01","마스터오더 생성","POST","/master-orders",
     "다수 오더 패킹. MO-YYYYMMDD-NNNNN 자동 생성",
     "-","-",
     "orderIds[](필수), serviceType(AIR|SEA|CIR|CCL,필수), remark",
     "masterOrderId, masterOrderNo, masterOrderStatus(CREATED), serviceType, orderCount",
     "INVALID_INPUT(400), ORDER_NOT_FOUND(404)"],
    ["API-W01","입고 등록","POST","/warehouse/receipts",
     "바코드 스캔/수동 입력. 오더 상태→WAREHOUSED",
     "-","-",
     "orderId(필수), barcode(필수), receivedQuantity(필수), receivedWeight, warehouseLocation, remark",
     "receiptId, orderId, orderNo, barcode, orderStatus(WAREHOUSED), receiptDate",
     "ORDER_NOT_FOUND(404), ORDER_ALREADY_WAREHOUSED(409)"],
    ["API-W04","출고 등록","POST","/warehouse/releases",
     "마스터오더 단위 출고 처리. 오더 상태→RELEASED",
     "-","-",
     "masterOrderId(필수), releaseType(NORMAL|URGENT), remark",
     "releaseId, masterOrderId, masterOrderNo, masterOrderStatus(RELEASED), releaseDate",
     "ORDER_NOT_FOUND(404)"],
    ["API-W07","바코드 스캔 조회","GET","/warehouse/barcode/{barcode}",
     "바코드로 오더 즉시 조회",
     "barcode(String,필수)","-","-",
     "orderId, orderNo, barcode, orderStatus, cargoName, quantity, consigneeName",
     "ORDER_NOT_FOUND(404)"],
    ["API-T01","트래킹 등록","POST","/tracking",
     "운송 상태 이력 등록. 갱신 임계값 30분",
     "-","-",
     "masterOrderId(필수), trackingStatus(필수), location, statusMessage, eventTime(필수), source(MANUAL|API_SYNC,필수)",
     "trackingId, masterOrderId, trackingStatus, eventTime",
     "ORDER_NOT_FOUND(404), INVALID_INPUT(400)"],
    ["API-T02","트래킹 이력 조회","GET","/tracking/{masterOrderId}",
     "전체 이력 최신순",
     "masterOrderId(Long,필수)","-","-",
     "masterOrderId, masterOrderNo, currentStatus, trackingHistory[]",
     "ORDER_NOT_FOUND(404)"],
    ["API-B01","청구서 생성","POST","/billing/invoices",
     "운송비 자동 계산. INV-YYYYMMDD-NNNNN 생성. 계산식: 원가+(원가×이익율)-(원가×(1-할인율))",
     "-","-",
     "masterOrderId(필수), extraCharge(소수점1자리), extraChargeNote, remark",
     "invoiceId, invoiceNo, masterOrderId, transportCost, extraCharge, totalAmount, invoiceStatus(UNPAID)",
     "ORDER_NOT_FOUND(404), INVALID_INPUT(400)"],
    ["API-B05","청구서 확정","PATCH","/billing/invoices/{invoiceId}/confirm",
     "확정 후 수정 불가",
     "invoiceId(Long,필수)","-","-",
     "invoiceId, invoiceStatus(CONFIRMED), confirmedAt",
     "INVOICE_ALREADY_CONFIRMED(409)"],
    ["API-B10","입금 등록","POST","/billing/payments",
     "전액/일부 입금. invoiceStatus 자동 갱신 (PARTIAL/PAID)",
     "-","-",
     "invoiceId(필수), paidAmount(필수), paymentMethod(BANK_TRANSFER|CARD|CASH), paymentDate(필수), memo",
     "paymentId, invoiceId, paidAmount, remainingAmount, invoiceStatus",
     "ORDER_NOT_FOUND(404)"],
    ["API-C10","운송비 계산 (미리보기)","POST","/costs/calculate",
     "오더 정보 기반 예상 운송비. 청구서 생성 전 확인용",
     "-","-",
     "orderId(필수), serviceType(필수), originCode(필수), destinationCode(필수), extraCharge",
     "transportCost, baseRate, profitAmount, discountAmount, extraCharge, totalAmount, calculation(설명문자열)",
     "ORDER_NOT_FOUND(404)"],
    ["API-N01","알림 목록","GET","/notifications",
     "본인 알림 목록. 읽음 여부 필터",
     "-",
     "page, size, isRead(boolean)","-",
     "content[], unreadCount, totalElements",
     "UNAUTHORIZED(401)"],
    ["API-V01","VOC 등록","POST","/voc",
     "민원/불만 접수",
     "-","-",
     "category(DELIVERY_DELAY|DAMAGE|LOST|ETC,필수), title(필수), content(필수), orderId, priority(LOW|MEDIUM|HIGH)",
     "vocId, title, vocStatus(OPEN), priority, createdAt",
     "INVALID_INPUT(400)"],
    ["API-V05","VOC 상태 변경","PATCH","/voc/{vocId}/status",
     "OPEN → IN_PROGRESS → CLOSED",
     "vocId(Long,필수)","-",
     "vocStatus(필수), comment",
     "vocId, vocStatus, updatedAt",
     "FORBIDDEN(403), NOT_FOUND(404)"],
    ["API-C02","운송원가 등록","POST","/costs/transport-masters",
     "서비스 유형/구간별 운송원가 등록",
     "-","-",
     "serviceType(필수), originCode(필수), destinationCode(필수), weightUnit, baseRate(필수), currency, profitRate, validFrom, validTo",
     "costId, serviceType, originCode, destinationCode, baseRate, profitRate",
     "INVALID_INPUT(400)"],
    ["API-S14","공통코드 목록","GET","/system/codes",
     "그룹별 공통코드. 프론트 드롭다운 용도",
     "-",
     "groupCode, useYn(Y|N)","-",
     "codeId, groupCode, codeValue, codeName, sortOrder, useYn",
     "UNAUTHORIZED(401)"],
    ["API-R01","국가코드 목록","GET","/codes/countries",
     "전체 국가코드 목록 (공항/항구/항공사/택배사 동일 구조)",
     "-",
     "keyword, useYn","-",
     "id, countryCode, countryNameKo, countryNameEn, useYn",
     "UNAUTHORIZED(401)"],
]

# ──────────────────────────────────────────────
# 헬퍼 함수
# ──────────────────────────────────────────────
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="맑은 고딕")

def make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def make_align(wrap=False, h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_header_row(ws, headers, row, fill_color, font_color="FFFFFF", bold=True):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = make_fill(fill_color)
        cell.font = make_font(bold=bold, color=font_color, size=10)
        cell.border = make_border()
        cell.alignment = make_align(h="center")

def set_data_cell(ws, row, col, value, fill_color=None, bold=False, wrap=False, h="left"):
    cell = ws.cell(row=row, column=col, value=value)
    if fill_color:
        cell.fill = make_fill(fill_color)
    cell.font = make_font(bold=bold, size=9)
    cell.border = make_border()
    cell.alignment = make_align(wrap=wrap, h=h)
    return cell

# ──────────────────────────────────────────────
# 1. SNTL_APIList.xlsx 생성
# ──────────────────────────────────────────────
def create_api_list_xlsx():
    wb = openpyxl.Workbook()

    # ── Sheet 1: API 목록
    ws = wb.active
    ws.title = "API목록"

    # 제목
    ws.merge_cells("A1:I1")
    title = ws["A1"]
    title.value = "SNTL 통합 물류 플랫폼 — API 목록 (Base URL: /api/v1)"
    title.fill = make_fill(COLOR["header_dark"])
    title.font = make_font(bold=True, color="FFFFFF", size=13)
    title.alignment = make_align(h="center")

    # 부제목
    ws.merge_cells("A2:I2")
    sub = ws["A2"]
    sub.value = f"총 {len(API_DATA)}개 API | 최종 업데이트: 2026-04-16"
    sub.fill = make_fill(COLOR["header_mid"])
    sub.font = make_font(bold=False, color="FFFFFF", size=10)
    sub.alignment = make_align(h="center")

    # 컬럼 헤더
    headers = ["API ID", "도메인", "Method", "URL", "기능명", "설명", "인증필요", "권한", "Phase"]
    set_header_row(ws, headers, 3, COLOR["header_light"], font_color="000000")

    # 컬럼 너비
    col_widths = [12, 14, 9, 55, 22, 42, 10, 15, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 틀 고정
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{3 + len(API_DATA)}"

    # 데이터 입력
    prev_domain = None
    for row_idx, row_data in enumerate(API_DATA, 4):
        api_id, domain, method, url, func_name, desc, auth, perm, phase = row_data

        domain_color = DOMAIN_COLOR.get(domain, COLOR["white"])
        method_color = METHOD_COLOR.get(method, COLOR["white"])
        phase_color  = PHASE_COLOR.get(phase, COLOR["white"])

        # 도메인 구분선 (다른 도메인 시작)
        if domain != prev_domain:
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="medium"),
                    bottom=Side(style="thin"),
                )
        prev_domain = domain

        set_data_cell(ws, row_idx, 1, api_id,    domain_color, bold=True, h="center")
        set_data_cell(ws, row_idx, 2, domain,     domain_color, h="center")
        # Method 셀에 색상 강조
        mc = ws.cell(row=row_idx, column=3, value=method)
        mc.fill  = make_fill(method_color)
        mc.font  = make_font(bold=True, color="FFFFFF", size=9)
        mc.border = make_border()
        mc.alignment = make_align(h="center")
        set_data_cell(ws, row_idx, 4, url,        None,  wrap=False)
        set_data_cell(ws, row_idx, 5, func_name,  None)
        set_data_cell(ws, row_idx, 6, desc,       None,  wrap=True)
        # 인증
        auth_fill = "C6EFCE" if auth == "Y" else "FFEB9C"
        set_data_cell(ws, row_idx, 7, auth, auth_fill, h="center")
        set_data_cell(ws, row_idx, 8, perm,       None,  h="center")
        # Phase
        set_data_cell(ws, row_idx, 9, f"Phase {phase}", phase_color, h="center")

    # ── Sheet 2: 도메인별 통계
    ws2 = wb.create_sheet("도메인통계")

    ws2.merge_cells("A1:D1")
    t2 = ws2["A1"]
    t2.value = "도메인별 API 통계"
    t2.fill = make_fill(COLOR["header_dark"])
    t2.font = make_font(bold=True, color="FFFFFF", size=12)
    t2.alignment = make_align(h="center")

    set_header_row(ws2, ["도메인", "API 수", "Phase", "비고"], 2, COLOR["header_light"], font_color="000000")

    domain_stat = {}
    for row in API_DATA:
        d = row[1]
        p = row[8]
        if d not in domain_stat:
            domain_stat[d] = {"count": 0, "phase": p}
        domain_stat[d]["count"] += 1

    domain_notes = {
        "인증": "JWT, SMS 인증",
        "회원-개인": "Soft Delete, 등급 BASIC~VIP",
        "회원-법인": "MinIO 첨부, 관리자 심사",
        "회원-담당자": "법인 소속 담당자",
        "회원-부서": "부서 + 부서담당자",
        "회원-등급": "등급별 정책 관리",
        "오더": "ORD-YYYYMMDD-NNNNN, AIR/SEA/CIR/CCL",
        "마스터오더": "MO-YYYYMMDD-NNNNN, 패킹",
        "창고": "바코드 스캔, 입출고",
        "Tracking": "갱신 임계 30분, API 타임아웃 10초",
        "회계/청구": "INV-YYYYMMDD-NNNNN, 운송비 계산",
        "원가": "운송원가, 할인율, 운송비 계산",
        "알림": "읽음/전체읽음/삭제",
        "VOC": "OPEN→IN_PROGRESS→CLOSED",
        "시스템": "메뉴/역할/코드그룹/공통코드",
        "기준코드": "국가/공항/항구/항공사/택배사",
    }

    for r_idx, (domain, stat) in enumerate(domain_stat.items(), 3):
        dc = DOMAIN_COLOR.get(domain, COLOR["white"])
        set_data_cell(ws2, r_idx, 1, domain, dc)
        set_data_cell(ws2, r_idx, 2, stat["count"], dc, h="center")
        set_data_cell(ws2, r_idx, 3, f"Phase {stat['phase']}", PHASE_COLOR.get(stat['phase'], COLOR["white"]), h="center")
        set_data_cell(ws2, r_idx, 4, domain_notes.get(domain, ""), None, wrap=True)

    # 합계
    total_row = 3 + len(domain_stat)
    ws2.cell(row=total_row, column=1, value="합계").font = make_font(bold=True)
    total_cell = ws2.cell(row=total_row, column=2, value=len(API_DATA))
    total_cell.font = make_font(bold=True)
    total_cell.fill = make_fill(COLOR["header_light"])
    total_cell.border = make_border()
    total_cell.alignment = make_align(h="center")
    for c in [1, 3, 4]:
        ws2.cell(row=total_row, column=c).border = make_border()
        ws2.cell(row=total_row, column=c).fill = make_fill(COLOR["header_light"])

    for i, w in enumerate([20, 8, 10, 35], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: 권한 범례
    ws3 = wb.create_sheet("권한범례")
    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = "권한 코드 범례"
    t3.fill = make_fill(COLOR["header_dark"])
    t3.font = make_font(bold=True, color="FFFFFF", size=12)
    t3.alignment = make_align(h="center")

    set_header_row(ws3, ["권한코드", "설명", "해당 역할"], 2, COLOR["header_light"], font_color="000000")

    perm_data = [
        ["PUBLIC",      "인증 불필요",                        "모든 사용자 (비로그인 포함)"],
        ["AUTH",        "인증된 모든 사용자",                  "ADMIN, OPERATOR, INDIVIDUAL, CORPORATE, DEPT_MANAGER"],
        ["ADMIN",       "관리자",                              "ADMIN"],
        ["ADMIN/OPR",   "관리자 또는 운영자",                  "ADMIN, OPERATOR"],
        ["SELF",        "본인 (또는 관리자)",                  "리소스 소유자 + ADMIN"],
        ["CORP",        "법인 관련 사용자",                    "법인관리자(CORPORATE), 부서담당자(DEPT_MANAGER), OPERATOR, ADMIN"],
        ["CORP_ADMIN",  "법인관리자 (또는 관리자)",            "CORPORATE, ADMIN"],
    ]

    for r_idx, row in enumerate(perm_data, 3):
        set_data_cell(ws3, r_idx, 1, row[0], COLOR["header_light"], bold=True, h="center")
        set_data_cell(ws3, r_idx, 2, row[1])
        set_data_cell(ws3, r_idx, 3, row[2], wrap=True)

    for i, w in enumerate([14, 26, 55], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save("SNTL_APIList.xlsx")
    print("SNTL_APIList.xlsx 생성 완료")


# ──────────────────────────────────────────────
# 2. SNTL_APISpec.xlsx 생성
# ──────────────────────────────────────────────
def create_api_spec_xlsx():
    wb = openpyxl.Workbook()

    # 표지 시트
    ws_cover = wb.active
    ws_cover.title = "표지"
    ws_cover.column_dimensions["A"].width = 80
    ws_cover.row_dimensions[3].height = 40
    ws_cover.row_dimensions[4].height = 22
    ws_cover.row_dimensions[5].height = 22

    ws_cover.merge_cells("A1:A1")
    c1 = ws_cover["A1"]
    c1.value = "SNTL 통합 물류 플랫폼"
    c1.fill = make_fill(COLOR["header_dark"])
    c1.font = Font(name="맑은 고딕", bold=True, size=20, color="FFFFFF")
    c1.alignment = make_align(h="center")

    c2 = ws_cover["A2"]
    c2.value = "API 정의서 (API Specification)"
    c2.fill = make_fill(COLOR["header_mid"])
    c2.font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")
    c2.alignment = make_align(h="center")

    for i, (k, v) in enumerate([
        ("Base URL", "http://localhost:8080/api/v1"),
        ("인증 방식", "JWT Bearer Token (Authorization: Bearer {access_token})"),
        ("응답 형식", "ApiResponse<T> 래퍼 { success, data, message }"),
        ("날짜 형식", "ISO 8601 (yyyy-MM-ddTHH:mm:ss)"),
        ("작성일", "2026-04-16"),
        ("총 API 수", f"{len(API_DATA)}개"),
    ], 4):
        ws_cover.cell(row=i, column=1, value=f"{k}: {v}").font = make_font(size=11)

    # ── 도메인별 시트
    domain_order = [
        "인증", "회원-개인", "회원-법인", "회원-담당자", "회원-부서", "회원-등급",
        "오더", "마스터오더", "창고", "Tracking",
        "회계/청구", "원가", "알림", "VOC", "시스템", "기준코드"
    ]

    # 도메인별 API 그룹핑
    domain_apis = {}
    for row in API_DATA:
        d = row[1]
        if d not in domain_apis:
            domain_apis[d] = []
        domain_apis[d].append(row)

    SPEC_HEADERS = ["API ID", "기능명", "Method", "URL", "설명",
                    "Path Params", "Query Params", "Request Body (주요 필드)",
                    "Response 200 (주요 필드)", "주요 에러코드"]
    COL_WIDTHS = [12, 20, 9, 50, 40, 30, 30, 50, 50, 45]

    # spec_dict: API_ID → spec 데이터
    spec_dict = {row[0]: row for row in API_SPEC_DATA}

    for domain in domain_order:
        if domain not in domain_apis:
            continue

        # 시트 이름 제한 (31자, 특수문자 제거)
        sheet_name = domain.replace("/", "_").replace("\\", "_").replace("*", "").replace("?", "").replace("[", "").replace("]", "")[:31]
        ws = wb.create_sheet(sheet_name)

        # 도메인 헤더
        ws.merge_cells(f"A1:{get_column_letter(len(SPEC_HEADERS))}1")
        dh = ws["A1"]
        dh.value = f"{domain} — API 정의서"
        dh.fill = make_fill(DOMAIN_COLOR.get(domain, COLOR["header_light"]))
        dh.font = make_font(bold=True, size=12)
        dh.alignment = make_align(h="center")

        set_header_row(ws, SPEC_HEADERS, 2, COLOR["header_light"], font_color="000000")

        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A3"

        row_idx = 3
        for api_row in domain_apis[domain]:
            api_id = api_row[0]
            method = api_row[2]

            spec = spec_dict.get(api_id)
            if spec:
                # spec 데이터 있음
                _, fname, sm, surl, sdesc, spath, squery, sreqbody, sresbody, serror = spec
            else:
                # spec 없으면 API_DATA 기본값 사용
                fname  = api_row[4]
                sm     = api_row[2]
                surl   = api_row[3]
                sdesc  = api_row[5]
                spath  = "-"
                squery = "-" if method in ["POST","PUT","DELETE","PATCH"] else "page, size, sort, keyword"
                sreqbody = "-" if method == "GET" else "(상세 정의서 참조)"
                sresbody = "data 객체 반환"
                serror   = "NOT_FOUND(404), FORBIDDEN(403)"

            method_color = METHOD_COLOR.get(method, COLOR["white"])
            row_fill = COLOR["light_gray"] if row_idx % 2 == 0 else COLOR["white"]

            set_data_cell(ws, row_idx, 1, api_id,    DOMAIN_COLOR.get(domain, row_fill), bold=True, h="center")
            set_data_cell(ws, row_idx, 2, fname,     row_fill)
            mc = ws.cell(row=row_idx, column=3, value=method)
            mc.fill  = make_fill(method_color)
            mc.font  = make_font(bold=True, color="FFFFFF", size=9)
            mc.border = make_border()
            mc.alignment = make_align(h="center")
            set_data_cell(ws, row_idx, 4, surl,      row_fill, wrap=True)
            set_data_cell(ws, row_idx, 5, sdesc,     row_fill, wrap=True)
            set_data_cell(ws, row_idx, 6, spath,     row_fill, wrap=True)
            set_data_cell(ws, row_idx, 7, squery,    row_fill, wrap=True)
            set_data_cell(ws, row_idx, 8, sreqbody,  row_fill, wrap=True)
            set_data_cell(ws, row_idx, 9, sresbody,  row_fill, wrap=True)
            set_data_cell(ws, row_idx, 10, serror,   row_fill, wrap=True)

            ws.row_dimensions[row_idx].height = 55
            row_idx += 1

    # ── 에러코드 범례 시트
    ws_err = wb.create_sheet("에러코드")
    ws_err.merge_cells("A1:D1")
    eh = ws_err["A1"]
    eh.value = "공통 에러 코드 목록"
    eh.fill = make_fill(COLOR["header_dark"])
    eh.font = make_font(bold=True, color="FFFFFF", size=12)
    eh.alignment = make_align(h="center")

    set_header_row(ws_err, ["ErrorCode", "HTTP Status", "메시지", "발생 상황"], 2, COLOR["header_light"], font_color="000000")

    error_data = [
        ["UNAUTHORIZED",             401, "인증이 필요합니다",                    "Bearer Token 없이 인증 필요 API 호출"],
        ["TOKEN_EXPIRED",            401, "토큰이 만료되었습니다",                "Access Token 30분 초과"],
        ["TOKEN_INVALID",            401, "유효하지 않은 토큰입니다",             "위조 또는 변조된 토큰"],
        ["FORBIDDEN",                403, "접근 권한이 없습니다",                 "타인 리소스 접근 또는 권한 부족"],
        ["NOT_FOUND",                404, "리소스를 찾을 수 없습니다",            "ID로 조회 실패"],
        ["MEMBER_NOT_FOUND",         404, "회원 정보를 찾을 수 없습니다",         "존재하지 않는 memberId"],
        ["ORDER_NOT_FOUND",          404, "오더 정보를 찾을 수 없습니다",         "존재하지 않는 orderId"],
        ["INVALID_INPUT",            400, "입력값이 올바르지 않습니다",           "필수 필드 누락 또는 형식 오류"],
        ["DUPLICATE_LOGIN_ID",       409, "이미 사용 중인 아이디입니다",          "loginId 중복"],
        ["PASSWORD_MISMATCH",        400, "비밀번호가 일치하지 않습니다",         "로그인 또는 탈퇴 시 비밀번호 불일치"],
        ["ACCOUNT_LOCKED",           423, "계정이 잠겨 있습니다",                "로그인 5회 실패 → 30분 잠금"],
        ["SMS_CODE_EXPIRED",         400, "SMS 인증번호가 만료되었습니다",        "3분 초과"],
        ["SMS_CODE_INVALID",         400, "SMS 인증번호가 올바르지 않습니다",     "인증번호 불일치"],
        ["CORPORATE_NOT_APPROVED",   403, "승인되지 않은 법인 회원입니다",        "approval_status=PENDING 상태에서 서비스 이용"],
        ["ORDER_ALREADY_WAREHOUSED", 409, "입고된 오더는 수정/삭제할 수 없습니다","orderStatus=WAREHOUSED 이상에서 수정/삭제 시도"],
        ["INVOICE_ALREADY_CONFIRMED",409, "이미 확정된 청구서입니다",            "invoiceStatus=CONFIRMED 이후 수정 시도"],
        ["INTERNAL_SERVER_ERROR",    500, "서버 내부 오류가 발생했습니다",        "예상치 못한 서버 오류"],
    ]

    for r_idx, row in enumerate(error_data, 3):
        bg = COLOR["light_gray"] if r_idx % 2 == 0 else COLOR["white"]
        set_data_cell(ws_err, r_idx, 1, row[0], bg, bold=True)
        set_data_cell(ws_err, r_idx, 2, row[1], bg, h="center")
        set_data_cell(ws_err, r_idx, 3, row[2], bg)
        set_data_cell(ws_err, r_idx, 4, row[3], bg, wrap=True)

    for i, w in enumerate([28, 12, 32, 50], 1):
        ws_err.column_dimensions[get_column_letter(i)].width = w

    wb.save("SNTL_APISpec.xlsx")
    print("SNTL_APISpec.xlsx 생성 완료")


# ──────────────────────────────────────────────
# 실행
# ──────────────────────────────────────────────
if __name__ == "__main__":
    create_api_list_xlsx()
    create_api_spec_xlsx()
    print(f"\n총 {len(API_DATA)}개 API 처리 완료")
    print(f"  - SNTL_APIList.xlsx : API 목록 + 도메인통계 + 권한범례")
    print(f"  - SNTL_APISpec.xlsx : 도메인별 상세정의 + 에러코드")
